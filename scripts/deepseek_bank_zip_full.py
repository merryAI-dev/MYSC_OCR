#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
import shutil
import subprocess
import sys
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.core import (  # noqa: E402
    AccountResult,
    classify_account_candidates,
    decode_zip_name,
    extract_roster,
    normalize_text,
    safe_filename_part,
    write_csv,
)


@dataclass(frozen=True)
class ZipMember:
    source_name: str
    size: int
    date_time: tuple[int, int, int, int, int, int]


def read_form_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = normalize_text(row[1] if len(row) > 1 else "")
        if not name or name == "테스트":
            continue
        rows.append(
            {
                "form_row": str(index),
                "timestamp": normalize_text(row[0] if len(row) > 0 else ""),
                "name": name,
                "phone": normalize_text(row[2] if len(row) > 2 else ""),
                "role": normalize_text(row[3] if len(row) > 3 else ""),
                "bank_link": normalize_text(row[6] if len(row) > 6 else ""),
            }
        )
    return rows


def zip_members(path: Path) -> list[ZipMember]:
    rows = []
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            decoded = decode_zip_name(info.filename)
            if info.is_dir() or Path(decoded).name == ".DS_Store":
                continue
            rows.append(ZipMember(decoded, info.file_size, info.date_time))
    return rows


def extract_member(zip_path: Path, member: ZipMember, destination: Path) -> Path:
    target = normalize_text(member.source_name)
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if decode_zip_name(info.filename) == target:
                destination.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(info) as source, destination.open("wb") as dest:
                    shutil.copyfileobj(source, dest)
                return destination
    raise FileNotFoundError(member.source_name)


def run(command: list[str], timeout: int = 120) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, text=True, capture_output=True, check=False, timeout=timeout)


def render_for_api(path: Path, output_dir: Path) -> list[Path]:
    if path.suffix.lower() != ".pdf":
        return [path]
    if not shutil.which("pdftoppm"):
        raise RuntimeError("pdftoppm is required for PDF bankbook OCR")
    prefix = output_dir / path.stem
    proc = run(["pdftoppm", "-png", "-r", "220", str(path), str(prefix)], timeout=120)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "pdftoppm failed")
    return sorted(output_dir.glob(f"{path.stem}-*.png"))


def deepseek_ocr(image_path: Path, api_url: str, timeout: int) -> str:
    prompt = (
        "<image>\n"
        "OCR this Korean bank account image. Copy only visible text and numbers. "
        "Focus on 계좌번호, 예금주, 은행명. "
        "Return lines like: bank: ..., account_holder: ..., account_number: ... "
        "If unknown, write UNKNOWN. Do not explain."
    )
    mime = mimetypes.guess_type(image_path.name)[0] or "image/png"
    with image_path.open("rb") as handle:
        response = requests.post(
            f"{api_url.rstrip('/')}/api/ocr",
            files={"file": (image_path.name, handle, mime)},
            data={
                "content_type": "Scene",
                "subcategory": "Verification",
                "complexity": "Tiny",
                "mode": "basic",
                "prompt": prompt,
            },
            timeout=timeout,
        )
    response.raise_for_status()
    data = response.json()
    if not data.get("success"):
        raise RuntimeError(data.get("error") or "DeepSeek OCR failed")
    return data.get("text", "")


def compact(value: str) -> str:
    return re.sub(r"[\s_()\\[\\]\\-]+", "", normalize_text(value)).lower()


def basename_hint(source_name: str) -> str:
    stem = Path(source_name).stem
    if " - " in stem:
        return normalize_text(stem.rsplit(" - ", 1)[-1])
    return normalize_text(stem)


def score_name(name: str, member: ZipMember, ocr_text: str) -> int:
    c_name = compact(name)
    source = compact(Path(member.source_name).name)
    hint = compact(basename_hint(member.source_name))
    ocr = compact(ocr_text)
    score = 0
    if c_name and c_name in source:
        score += 80
    if c_name and c_name in hint:
        score += 100
    if c_name and c_name in ocr:
        score += 80
    return score


def best_name_match(names: list[str], member: ZipMember, ocr_text: str) -> tuple[str, int, str]:
    scored = sorted(
        ((name, score_name(name, member, ocr_text)) for name in names),
        key=lambda item: item[1],
        reverse=True,
    )
    if not scored or scored[0][1] <= 0:
        return "", 0, "no_name_signal"
    if len(scored) > 1 and scored[1][1] == scored[0][1]:
        return scored[0][0], scored[0][1], "ambiguous_name_score"
    return scored[0][0], scored[0][1], "filename_or_ocr_name_match"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--form-xlsx", type=Path, required=True)
    parser.add_argument("--bank-zip", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--api-url", default="http://127.0.0.1:5001")
    parser.add_argument("--timeout", type=int, default=900)
    parser.add_argument("--limit", type=int, default=0)
    args = parser.parse_args()

    health = requests.get(f"{args.api_url.rstrip('/')}/api/health", timeout=10)
    health.raise_for_status()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    extracted_dir = args.output_dir / "bank_zip_extracted"
    rendered_dir = args.output_dir / "rendered"
    ocr_text_dir = args.output_dir / "deepseek_text"
    extracted_dir.mkdir(parents=True, exist_ok=True)
    rendered_dir.mkdir(parents=True, exist_ok=True)
    ocr_text_dir.mkdir(parents=True, exist_ok=True)

    roster = extract_roster(args.workbook)
    form_rows = read_form_rows(args.form_xlsx)
    name_pool = sorted(set(roster.names) | {row["name"] for row in form_rows})

    rows = []
    members = zip_members(args.bank_zip)
    if args.limit:
        members = members[: args.limit]

    for index, member in enumerate(members, start=1):
        print(f"DeepSeek OCR {index}/{len(members)} {Path(member.source_name).name}", flush=True)
        safe_stem = f"{index:03d}_{safe_filename_part(Path(member.source_name).stem)}"
        local_path = extracted_dir / f"{safe_stem}{Path(member.source_name).suffix}"
        extract_member(args.bank_zip, member, local_path)
        page_paths = render_for_api(local_path, rendered_dir)

        page_texts = []
        error = ""
        started = time.time()
        try:
            for page_path in page_paths[:1]:
                page_texts.append(deepseek_ocr(page_path, args.api_url, args.timeout))
        except Exception as exc:
            error = f"{type(exc).__name__}: {exc}"
        elapsed = time.time() - started
        text = "\n".join(page_texts)
        text_path = ocr_text_dir / f"{safe_stem}.txt"
        text_path.write_text(text, encoding="utf-8")

        account_result = classify_account_candidates(text)
        if error:
            account_result = AccountResult(None, "error", [], error, "mlx-deepseek")

        matched_name, name_score, match_reason = best_name_match(name_pool, member, text)
        form_match = next((row for row in form_rows if row["name"] == matched_name), None)
        roster_match = next((person for person in roster.people if person.name == matched_name), None)
        dt = member.date_time
        rows.append(
            {
                "index": index,
                "zip_datetime": f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}",
                "source_name": member.source_name,
                "extracted_path": str(local_path),
                "filename_hint": basename_hint(member.source_name),
                "matched_name": matched_name,
                "name_score": name_score,
                "match_reason": match_reason,
                "form_row": form_match["form_row"] if form_match else "",
                "form_role": form_match["role"] if form_match else "",
                "roster_group": roster_match.group if roster_match else "",
                "roster_no": roster_match.no if roster_match else "",
                "account": account_result.value or "",
                "confidence": account_result.confidence,
                "candidates": "; ".join(account_result.candidates),
                "account_reason": account_result.reason,
                "elapsed_seconds": f"{elapsed:.2f}",
                "ocr_text_path": str(text_path),
            }
        )
        write_csv(args.output_dir / "deepseek_bank_zip_full_ocr.csv", rows, CSV_FIELDS)

    summary = {
        "members": len(rows),
        "high_accounts": sum(1 for row in rows if row["confidence"] == "high"),
        "matched_names": sum(1 for row in rows if row["matched_name"]),
        "output": str(args.output_dir),
    }
    (args.output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


CSV_FIELDS = [
    "index",
    "zip_datetime",
    "source_name",
    "extracted_path",
    "filename_hint",
    "matched_name",
    "name_score",
    "match_reason",
    "form_row",
    "form_role",
    "roster_group",
    "roster_no",
    "account",
    "confidence",
    "candidates",
    "account_reason",
    "elapsed_seconds",
    "ocr_text_path",
]


if __name__ == "__main__":
    raise SystemExit(main())
