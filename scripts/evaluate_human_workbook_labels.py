#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.core import extract_roster, normalize_text, write_csv  # noqa: E402


DETAIL_FIELDS = [
    "group",
    "no",
    "name",
    "human_status",
    "outcome",
    "human_bank",
    "human_account_masked",
    "predicted_account_masked",
    "decision",
    "source",
    "candidate_files",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    return list(csv.DictReader(path.open(encoding="utf-8-sig")))


def digits_only(value: object) -> str:
    return re.sub(r"\D+", "", normalize_text(value))


def mask_account_for_report(value: object, *, keep_last: int = 3) -> str:
    text = normalize_text(value)
    digits = [index for index, char in enumerate(text) if char.isdigit()]
    if not digits:
        return text
    keep = set(digits[-keep_last:])
    return "".join(char if index in keep or not char.isdigit() else "*" for index, char in enumerate(text))


def load_review_names(workbook_path: Path) -> set[str]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    if "REMAINING_REVIEW" not in workbook.sheetnames:
        return set()
    sheet = workbook["REMAINING_REVIEW"]
    headers = [normalize_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    try:
        name_column = headers.index("name") + 1
    except ValueError:
        name_column = 3
    names = set()
    for row in range(2, sheet.max_row + 1):
        name = normalize_text(sheet.cell(row, name_column).value)
        if name:
            names.add(name)
    return names


def load_human_labels(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    sheet = workbook["시트"] if "시트" in workbook.sheetnames else workbook.active
    review_names = load_review_names(workbook_path)
    rows = []
    for person in extract_roster(workbook_path).people:
        account = normalize_text(sheet.cell(person.row, 11).value)
        bank = normalize_text(sheet.cell(person.row, 10).value)
        status = "review" if person.name in review_names else "positive"
        rows.append(
            {
                "group": person.group,
                "no": person.no,
                "name": person.name,
                "status": status,
                "bank": bank,
                "account_digits": digits_only(account),
                "account_masked": mask_account_for_report(account) if account else "",
            }
        )
    return rows


def load_predictions(resolution_csv: Path) -> dict[str, dict[str, str]]:
    predictions = {}
    for row in read_csv(resolution_csv):
        account = row.get("chosen_account", "")
        masked_source = account or row.get("chosen_account_masked", "")
        predictions[row.get("name", "")] = {
            "decision": row.get("decision", ""),
            "source": row.get("source", ""),
            "candidate_files": row.get("candidate_files", ""),
            "account_digits": digits_only(account),
            "account_masked": mask_account_for_report(masked_source) if masked_source else "",
        }
    return predictions


def classify_outcome(label: dict[str, Any], prediction: dict[str, str]) -> str:
    predicted_digits = prediction.get("account_digits", "")
    if label["status"] == "review":
        return "review_false_positive" if predicted_digits else "review_deferred"
    if predicted_digits and predicted_digits == label["account_digits"]:
        return "correct_positive"
    if predicted_digits:
        return "wrong_positive"
    return "missed_positive"


def zip_metadata(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    with zipfile.ZipFile(path) as archive:
        infos = [info for info in archive.infolist() if not info.is_dir() and Path(info.filename).name != ".DS_Store"]
    suffix_counts: dict[str, int] = {}
    for info in infos:
        suffix = Path(info.filename).suffix.lower() or "<none>"
        suffix_counts[suffix] = suffix_counts.get(suffix, 0) + 1
    return {
        "path": str(path),
        "file_count": len(infos),
        "total_uncompressed_mb": round(sum(info.file_size for info in infos) / 1024 / 1024, 1),
        "suffix_counts": dict(sorted(suffix_counts.items())),
    }


def _increment(stats: dict[str, int], key: str) -> None:
    stats[key] = int(stats.get(key, 0)) + 1


def evaluate_human_workbook_labels(
    *,
    human_workbook: Path,
    resolution_csv: Path,
    output_dir: Path,
    data_zip: Path | None = None,
) -> dict[str, Any]:
    labels = load_human_labels(human_workbook)
    predictions = load_predictions(resolution_csv)
    detail_rows = []
    summary: dict[str, int] = {
        "human_positive_count": 0,
        "human_review_count": 0,
        "correct_positive": 0,
        "wrong_positive": 0,
        "missed_positive": 0,
        "review_false_positive": 0,
        "review_deferred": 0,
    }
    by_decision: dict[str, dict[str, int]] = defaultdict(dict)

    for label in labels:
        prediction = predictions.get(label["name"], {})
        decision = prediction.get("decision", "missing_prediction_row")
        outcome = classify_outcome(label, prediction)
        if label["status"] == "positive":
            _increment(summary, "human_positive_count")
        else:
            _increment(summary, "human_review_count")
        _increment(summary, outcome)
        _increment(by_decision[decision], outcome)
        _increment(by_decision[decision], "total")
        detail_rows.append(
            {
                "group": label["group"],
                "no": label["no"],
                "name": label["name"],
                "human_status": label["status"],
                "outcome": outcome,
                "human_bank": label["bank"],
                "human_account_masked": label["account_masked"] if label["status"] == "positive" else "",
                "predicted_account_masked": prediction.get("account_masked", ""),
                "decision": decision,
                "source": prediction.get("source", ""),
                "candidate_files": prediction.get("candidate_files", ""),
            }
        )

    selected_positive = summary["correct_positive"] + summary["wrong_positive"]
    positive_total = summary["human_positive_count"]
    selected_total = selected_positive + summary["review_false_positive"]
    metrics = {
        "positive_precision": summary["correct_positive"] / selected_positive if selected_positive else 0.0,
        "safe_selection_precision": summary["correct_positive"] / selected_total if selected_total else 0.0,
        "positive_recall": summary["correct_positive"] / positive_total if positive_total else 0.0,
        "review_false_positive_rate": summary["review_false_positive"] / summary["human_review_count"]
        if summary["human_review_count"]
        else 0.0,
    }
    report = {
        "human_workbook": str(human_workbook),
        "resolution_csv": str(resolution_csv),
        "data_zip": zip_metadata(data_zip),
        "summary": summary,
        "metrics": metrics,
        "by_decision": dict(sorted(by_decision.items())),
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "human_label_eval_details.csv", detail_rows, DETAIL_FIELDS)
    (output_dir / "human_label_eval.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate a masked resolution CSV against a human-labeled workbook.")
    parser.add_argument("--human-workbook", type=Path, required=True)
    parser.add_argument("--resolution-csv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--data-zip", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    report = evaluate_human_workbook_labels(
        human_workbook=args.human_workbook,
        resolution_csv=args.resolution_csv,
        output_dir=args.output_dir,
        data_zip=args.data_zip,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
