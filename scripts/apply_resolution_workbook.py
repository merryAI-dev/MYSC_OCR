#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook

import sys

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.core import extract_roster, write_csv  # noqa: E402
from scripts.evaluate_human_workbook_labels import load_review_names  # noqa: E402


WORKBOOK_AUTO_FILL_DECISIONS = {
    "keep_existing_final_run",
    "auto_fill_single_candidate",
    "auto_fill_targeted_deepseek",
    "auto_fill_policy_reranker",
    "auto_fill_targeted_policy_reranker",
    "auto_fill_openai_reranker",
    "auto_fill_union_fallback",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    return list(csv.DictReader(path.open(encoding="utf-8-sig")))


def is_workbook_auto_fill_decision(decision: str) -> bool:
    return decision in WORKBOOK_AUTO_FILL_DECISIONS


def apply_resolution_workbook(
    *,
    source_workbook: Path,
    resolution_csv: Path,
    output_dir: Path,
    manual_review_workbook: Path | None = None,
) -> dict[str, object]:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_workbook = output_dir / f"{source_workbook.stem}_계좌번호입력_deepseek보강{source_workbook.suffix}"

    roster = extract_roster(source_workbook)
    by_name = {row["name"]: row for row in read_csv(resolution_csv)}
    hard_gate_names = load_review_names(manual_review_workbook) if manual_review_workbook else set()

    wb = load_workbook(source_workbook, read_only=False, data_only=False)
    ws = wb.active
    report = []

    for person in roster.people:
        row = by_name.get(person.name, {})
        account = row.get("chosen_account", "")
        decision = row.get("decision", "missing_resolution")
        cell = ws[f"J{person.row}"]
        hard_gate_reason = ""
        if person.name in hard_gate_names:
            status = "hard_gate_blocked"
            hard_gate_reason = "manual_review_required"
        elif account and is_workbook_auto_fill_decision(decision):
            cell.value = account
            cell.number_format = "@"
            status = "updated"
        else:
            status = "skipped"
        report.append(
            {
                "group": person.group,
                "no": person.no,
                "name": person.name,
                "cell": cell.coordinate,
                "status": status,
                "decision": decision,
                "source": row.get("source", ""),
                "account": account,
                "account_masked": row.get("chosen_account_masked", ""),
                "candidate_count": row.get("candidate_count", ""),
                "candidate_accounts_masked": row.get("candidate_accounts_masked", ""),
                "candidate_files": row.get("candidate_files", ""),
                "hard_gate_reason": hard_gate_reason,
            }
        )

    wb.save(output_workbook)
    write_csv(
        output_dir / "account_updates_deepseek_resolution.csv",
        report,
        [
            "group",
            "no",
            "name",
            "cell",
            "status",
            "decision",
            "source",
            "account",
            "account_masked",
            "candidate_count",
            "candidate_accounts_masked",
            "candidate_files",
            "hard_gate_reason",
        ],
    )
    summary = {
        "output_workbook": str(output_workbook),
        "updated": sum(1 for row in report if row["status"] == "updated"),
        "skipped": sum(1 for row in report if row["status"] == "skipped"),
        "hard_gate_blocked": sum(1 for row in report if row["status"] == "hard_gate_blocked"),
        "output_dir": str(output_dir),
    }
    (output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", type=Path, required=True)
    parser.add_argument("--resolution-csv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--manual-review-workbook", type=Path)
    args = parser.parse_args()

    summary = apply_resolution_workbook(
        source_workbook=args.source_workbook,
        resolution_csv=args.resolution_csv,
        output_dir=args.output_dir,
        manual_review_workbook=args.manual_review_workbook,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
