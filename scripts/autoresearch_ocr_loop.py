#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import os
import re
import signal
import shutil
import subprocess
import sys
import time
import urllib.error
import urllib.request
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from contextlib import contextmanager
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter, ImageOps

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.core import (  # noqa: E402
    AccountResult,
    classify_account_candidates,
    extract_roster,
    normalize_text,
    safe_filename_part,
    write_csv,
)
from settlement_tool.account_policy import policy_score_candidate  # noqa: E402
from settlement_tool.ocr import OcrOptions, extract_text_chandra  # noqa: E402
from scripts.build_deepseek_resolution import MANUAL_NAME_HINTS, SOURCE_NAME_HINTS, mask  # noqa: E402


PROMPTS = [
    {
        "id": "account_structured_ko_v2",
        "text": (
            "<image>\n"
            "한국 은행 통장사본 OCR입니다. 보이는 텍스트만 읽으세요. "
            "계좌번호/예금주/은행명을 찾고 추측하지 마세요. "
            "주민등록번호, 운전면허번호, 전화번호, 날짜는 계좌번호가 아닙니다. "
            "출력: bank: ...\\naccount_holder: ...\\naccount_number: ..."
        ),
    },
    {
        "id": "number_inventory_ko",
        "text": (
            "<image>\n"
            "이미지에서 보이는 숫자 묶음을 줄마다 그대로 복사하세요. "
            "각 숫자 옆에 주변 라벨을 함께 적으세요. 예: 계좌번호: 000-000-000000. "
            "설명하지 마세요."
        ),
    },
    {
        "id": "account_only_negative_rules",
        "text": (
            "<image>\n"
            "Find the Korean bank account number only. "
            "Reject phone numbers, dates, timestamps, resident IDs, driver license numbers, and card numbers. "
            "Return exactly one line: account_number: <visible number or UNKNOWN>."
        ),
    },
    {
        "id": "bankbook_copy_all_text",
        "text": (
            "<image>\n"
            "Copy all visible Korean/English text and numbers from this bankbook/account-copy image. "
            "Preserve line breaks when possible. Do not infer missing text."
        ),
    },
]

CHANDRA_PROMPTS = [
    (
        "chandra_bank_fields_v2",
        "Read this Korean bankbook copy. Extract only visible bank account information. "
        "Return bank, account_holder, account_number. Do not guess. "
        "Do not return resident ID, driver license, phone, date, or timestamp numbers as account numbers.",
    ),
    (
        "chandra_numbers_with_labels",
        "OCR this image. Copy every visible number group with nearby Korean label. "
        "Focus on labels such as 계좌번호, 입금계좌, 예금주, 은행, 통장. Do not explain.",
    ),
]


class OcrTimeoutError(TimeoutError):
    pass


@contextmanager
def time_limit(seconds: int, label: str):
    if seconds <= 0:
        yield
        return

    def handler(_signum, _frame):
        raise OcrTimeoutError(f"{label} timed out after {seconds}s")

    old_handler = signal.signal(signal.SIGALRM, handler)
    signal.alarm(seconds)
    try:
        yield
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)


FIELDS = [
    "name",
    "backend",
    "prompt_id",
    "variant",
    "source_name",
    "account",
    "account_masked",
    "confidence",
    "candidates",
    "candidates_masked",
    "reason",
    "failure_type",
    "elapsed_seconds",
    "text_path",
    "image_path",
    "error",
]


RESOLUTION_FIELDS = [
    "group",
    "no",
    "name",
    "prior_status",
    "prior_account_masked",
    "chosen_account",
    "chosen_account_masked",
    "decision",
    "source",
    "candidate_count",
    "candidate_accounts_masked",
    "candidate_files",
]


@dataclass(frozen=True)
class Evidence:
    name: str
    account: str
    backend: str
    prompt_id: str
    variant: str
    source_name: str
    confidence: str
    reason: str
    text_path: str


def read_csv(path: Path) -> list[dict[str, str]]:
    return list(csv.DictReader(path.open(encoding="utf-8-sig")))


def append_jsonl(path: Path, payload: dict[str, object]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


def run(command: list[str], timeout: int = 120) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, text=True, capture_output=True, check=False, timeout=timeout)


def render_pdf(path: Path, output_dir: Path) -> list[Path]:
    if path.suffix.lower() != ".pdf":
        return [path]
    if not shutil.which("pdftoppm"):
        raise RuntimeError("pdftoppm is required for PDF OCR")
    prefix = output_dir / safe_filename_part(path.stem)
    proc = run(["pdftoppm", "-png", "-r", "320", str(path), str(prefix)], timeout=180)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "pdftoppm failed")
    return sorted(output_dir.glob(f"{prefix.name}-*.png"))[:2]


def make_variants(image_path: Path, output_dir: Path, cycle: int) -> list[tuple[str, Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    variants: list[tuple[str, Path]] = [("original", image_path)]
    with Image.open(image_path) as raw:
        img = ImageOps.exif_transpose(raw).convert("RGB")
    width, height = img.size
    if max(width, height) < 1800:
        scale = 2
        img = img.resize((width * scale, height * scale), Image.Resampling.LANCZOS)
        width, height = img.size

    gray = ImageOps.grayscale(img)
    gray = ImageOps.autocontrast(gray)
    gray = ImageEnhance.Contrast(gray).enhance(1.6 + min(cycle % 4, 3) * 0.25)
    gray = gray.filter(ImageFilter.SHARPEN)
    contrast_path = output_dir / f"{safe_filename_part(image_path.stem)}__c{cycle}_contrast.png"
    gray.save(contrast_path)
    variants.append(("contrast", contrast_path))

    crops = [
        ("top_65", (0, 0, width, int(height * 0.65))),
        ("middle_70", (0, int(height * 0.15), width, int(height * 0.85))),
        ("left_70", (0, 0, int(width * 0.70), height)),
        ("right_70", (int(width * 0.30), 0, width, height)),
    ]
    for name, box in crops:
        crop = gray.crop(box)
        if crop.size[0] < 200 or crop.size[1] < 200:
            continue
        path = output_dir / f"{safe_filename_part(image_path.stem)}__c{cycle}_{name}.png"
        crop.save(path)
        variants.append((name, path))
    return variants


def deepseek_form_fields(
    prompt: str,
    *,
    max_tokens: int,
    early_stop_account: bool,
    prefix_salvage: bool,
    repetition_penalty: float,
    repetition_context_size: int,
) -> dict[str, str]:
    fields = {
        "content_type": "Scene",
        "subcategory": "Verification",
        "complexity": "Tiny",
        "prompt": prompt,
        "max_tokens": str(max_tokens),
        "early_stop_account": "1" if early_stop_account else "0",
        "prefix_salvage": "1" if prefix_salvage else "0",
    }
    if repetition_penalty > 0:
        fields["repetition_penalty"] = f"{repetition_penalty:.4g}"
        fields["repetition_context_size"] = str(repetition_context_size)
    return fields


def post_deepseek(
    image_path: Path,
    prompt: str,
    api_url: str,
    timeout: int,
    *,
    max_tokens: int,
    early_stop_account: bool,
    prefix_salvage: bool,
    repetition_penalty: float,
    repetition_context_size: int,
) -> str:
    boundary = f"----codex-{uuid.uuid4().hex}"
    mime = mimetypes.guess_type(image_path.name)[0] or "image/png"
    fields = deepseek_form_fields(
        prompt,
        max_tokens=max_tokens,
        early_stop_account=early_stop_account,
        prefix_salvage=prefix_salvage,
        repetition_penalty=repetition_penalty,
        repetition_context_size=repetition_context_size,
    )
    body = bytearray()
    for key, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode())
        body.extend(str(value).encode("utf-8"))
        body.extend(b"\r\n")
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        f'Content-Disposition: form-data; name="file"; filename="{image_path.name}"\r\n'
        f"Content-Type: {mime}\r\n\r\n".encode()
    )
    body.extend(image_path.read_bytes())
    body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode())

    request = urllib.request.Request(
        f"{api_url.rstrip('/')}/api/ocr",
        data=bytes(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if not payload.get("success"):
        raise RuntimeError(payload.get("error") or "DeepSeek OCR failed")
    return payload.get("text", "")


def is_deepseek_down_error(error: str) -> bool:
    lowered = error.lower()
    return (
        "connection refused" in lowered
        or "remote end closed connection" in lowered
        or "connection reset" in lowered
        or "broken pipe" in lowered
    )


def account_key(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def looks_like_id_or_timestamp(candidate: str, text: str) -> bool:
    digits = account_key(candidate)
    if len(digits) == 13:
        birth = digits[:6]
        marker = digits[6:7]
        if marker in {"1", "2", "3", "4", "5", "6"} and re.fullmatch(r"\d{6}", birth):
            return True
    if len(digits) >= 10 and digits.startswith("20"):
        return True
    window = text[max(0, text.find(candidate) - 50) : text.find(candidate) + len(candidate) + 50] if candidate in text else text
    id_words = ("주민", "운전", "면허", "등록번호", "license", "resident", "생년월일")
    return any(word.lower() in window.lower() for word in id_words)


def classify_with_labels(text: str) -> AccountResult:
    result = classify_account_candidates(text)
    if result.value:
        return result

    label_re = re.compile(
        r"(?:account[_ ]?number|계좌\s*번호|입금\s*계좌|계좌|Account)\D{0,30}((?:\d[\d -]{7,22}\d))",
        re.IGNORECASE,
    )
    candidates: list[str] = []
    for match in label_re.finditer(text):
        candidate = re.sub(r"[^\d-]", "", match.group(1)).strip("-")
        digits = account_key(candidate)
        decision = policy_score_candidate(text, candidate)
        if 9 <= len(digits) <= 16 and not looks_like_id_or_timestamp(candidate, text) and decision.accepted:
            candidates.append(candidate)
    seen = []
    for candidate in candidates:
        if candidate not in seen:
            seen.append(candidate)
    if len(seen) == 1:
        return AccountResult(seen[0], "high", seen, "single_account_label_match", "label_parser")
    if seen:
        return AccountResult(None, "low", seen, "multiple_account_label_matches", "label_parser")
    return result


def failure_type(result: AccountResult, text: str, error: str) -> str:
    if error:
        return "backend_error"
    if result.candidates and any(looks_like_id_or_timestamp(candidate, text) for candidate in result.candidates):
        return "privacy_id_like"
    if result.confidence == "low" and result.candidates:
        return "ambiguous_or_low_confidence"
    if not result.candidates:
        return "ocr_no_candidate"
    return ""


def target_name(row: dict[str, str], target_names: set[str]) -> str:
    source_manual = SOURCE_NAME_HINTS.get(Path(row.get("source_name", "")).stem, "")
    if source_manual in target_names:
        return source_manual
    if row.get("matched_name") in target_names:
        return row["matched_name"]
    manual = MANUAL_NAME_HINTS.get(row.get("filename_hint", ""), "")
    if manual in target_names:
        return manual
    source = row.get("source_name", "")
    for name in target_names:
        if name and name in source:
            return name
    return ""


def build_targets(resolution_rows: list[dict[str, str]], deepseek_rows: list[dict[str, str]]) -> tuple[list[tuple[str, dict[str, str]]], list[str]]:
    target_names = {row["name"] for row in resolution_rows if not row.get("chosen_account")}
    targets = []
    seen = set()
    for row in deepseek_rows:
        name = target_name(row, target_names)
        if not name:
            continue
        key = (name, row["source_name"])
        if key in seen:
            continue
        seen.add(key)
        targets.append((name, row))
    with_files = {name for name, _ in targets}
    missing = sorted(target_names - with_files)
    return targets, missing


def select_prompts(prompt_scores: Counter[str], cycle: int) -> list[dict[str, str]]:
    if cycle < 2:
        return PROMPTS
    ranked = sorted(PROMPTS, key=lambda item: (-prompt_scores[item["id"]], item["id"]))
    explore = PROMPTS[cycle % len(PROMPTS)]
    selected = ranked[:3]
    if explore not in selected:
        selected.append(explore)
    return selected


def choose_accounts(evidence: list[Evidence]) -> dict[str, tuple[str, str, list[Evidence]]]:
    by_name_key: dict[tuple[str, str], list[Evidence]] = defaultdict(list)
    for item in evidence:
        key = account_key(item.account)
        if key:
            by_name_key[(item.name, key)].append(item)

    by_name: dict[str, list[tuple[str, list[Evidence]]]] = defaultdict(list)
    for (name, _), items in by_name_key.items():
        by_name[name].append((items[0].account, items))

    chosen: dict[str, tuple[str, str, list[Evidence]]] = {}
    for name, candidates in by_name.items():
        qualified = []
        for account, items in candidates:
            backends = {item.backend for item in items}
            prompts = {item.prompt_id for item in items}
            variants = {item.variant for item in items}
            if len(backends) >= 2:
                qualified.append((account, "auto_fill_autoresearch_cross_backend", items))
            elif len(prompts) >= 2 and len(variants) >= 2 and len(items) >= 3:
                qualified.append((account, "auto_fill_autoresearch_single_backend_consensus", items))
        if len(qualified) == 1 and len(candidates) == 1:
            chosen[name] = qualified[0]
    return chosen


def write_resolution(
    base_rows: list[dict[str, str]],
    chosen: dict[str, tuple[str, str, list[Evidence]]],
    output_dir: Path,
) -> Path:
    rows = []
    for row in base_rows:
        current = dict(row)
        if not current.get("chosen_account") and current["name"] in chosen:
            account, decision, items = chosen[current["name"]]
            current["chosen_account"] = account
            current["chosen_account_masked"] = mask(account)
            current["decision"] = decision
            current["source"] = ";".join(sorted({f"{item.backend}:{item.prompt_id}:{item.variant}" for item in items}))
            current["candidate_count"] = "1"
            current["candidate_accounts_masked"] = mask(account)
            current["candidate_files"] = "; ".join(sorted({item.source_name for item in items}))
        rows.append(current)
    path = output_dir / "account_resolution_candidates.csv"
    write_csv(path, rows, RESOLUTION_FIELDS)
    return path


def write_workbook(source_workbook: Path, resolution_rows: list[dict[str, str]], output_dir: Path) -> Path:
    roster = extract_roster(source_workbook)
    by_name = {row["name"]: row for row in resolution_rows}
    wb = load_workbook(source_workbook, read_only=False, data_only=False)
    ws = wb.active
    report = []
    for person in roster.people:
        row = by_name.get(person.name, {})
        account = row.get("chosen_account", "")
        cell = ws[f"J{person.row}"]
        if account:
            cell.value = account
            cell.number_format = "@"
            status = "updated"
        else:
            status = "skipped"
        report.append(
            {
                "group": person.group,
                "no": person.no,
                "name": person.name,
                "cell": cell.coordinate,
                "status": status,
                "decision": row.get("decision", ""),
                "source": row.get("source", ""),
                "account": account,
                "account_masked": row.get("chosen_account_masked", ""),
            }
        )
    path = output_dir / f"{source_workbook.stem}_계좌번호입력_autoresearch{source_workbook.suffix}"
    wb.save(path)
    write_csv(output_dir / "account_updates_autoresearch.csv", report, list(report[0].keys()))
    return path


def write_program(output_dir: Path) -> None:
    (output_dir / "program.md").write_text(
        "\n".join(
            [
                "# Settlement OCR Autoresearch Program",
                "",
                "Goal: improve unresolved Korean bankbook OCR without sending sensitive files to remote APIs.",
                "",
                "Loop:",
                "1. Run DeepSeek MLX OCR with a ranked prompt library.",
                "2. Run Chandra local MPS OCR in the same Python process so the model is loaded once.",
                "3. Generate image variants: original, contrast, top/middle/left/right crops.",
                "4. Classify failures for prompt tuning and future fine-tuning.",
                "5. Auto-fill only when candidates converge under conservative consensus rules.",
                "",
                "Failure labels: missing_file, backend_error, ocr_no_candidate, ambiguous_or_low_confidence, privacy_id_like, conflicting_candidates.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-workbook", type=Path, required=True)
    parser.add_argument("--base-resolution-csv", type=Path, required=True)
    parser.add_argument("--deepseek-csv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--duration-hours", type=float, default=12)
    parser.add_argument("--api-url", default="http://127.0.0.1:5001")
    parser.add_argument("--timeout", type=int, default=900)
    parser.add_argument("--cycle-sleep-seconds", type=int, default=60)
    parser.add_argument("--chandra-device", default="mps")
    parser.add_argument("--chandra-model", default="models/chandra-ocr-2")
    parser.add_argument("--chandra-max-output-tokens", type=int, default=160)
    parser.add_argument("--chandra-image-max-side", type=int, default=1600)
    parser.add_argument("--chandra-timeout-seconds", type=int, default=120)
    parser.add_argument("--disable-chandra", action="store_true", help="Run only DeepSeek prompts and skip local Chandra OCR.")
    parser.add_argument("--deepseek-max-tokens", type=int, default=512)
    parser.add_argument("--deepseek-early-stop-account", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--deepseek-prefix-salvage", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--deepseek-repetition-penalty", type=float, default=1.05)
    parser.add_argument("--deepseek-repetition-context-size", type=int, default=64)
    parser.add_argument("--max-cycles", type=int, default=0)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    rendered_dir = args.output_dir / "rendered"
    variants_dir = args.output_dir / "variants"
    text_dir = args.output_dir / "ocr_text"
    for path in (rendered_dir, variants_dir, text_dir):
        path.mkdir(parents=True, exist_ok=True)
    write_program(args.output_dir)

    base_rows = read_csv(args.base_resolution_csv)
    deepseek_rows = read_csv(args.deepseek_csv)
    targets, missing_files = build_targets(base_rows, deepseek_rows)
    for name in missing_files:
        append_jsonl(
            args.output_dir / "failure_corpus.jsonl",
            {"name": name, "failure_type": "missing_file", "source_name": "", "note": "No matching bankbook file in ZIP"},
        )

    deadline = time.time() + args.duration_hours * 3600
    prompt_scores: Counter[str] = Counter()
    all_rows: list[dict[str, str]] = []
    evidence: list[Evidence] = []
    best_chosen_count = 0
    cycle = 0
    chandra_disabled_reason = "disabled_by_flag" if args.disable_chandra else ""

    print(json.dumps({"event": "start", "targets": len(targets), "missing_files": missing_files, "output_dir": str(args.output_dir)}, ensure_ascii=False), flush=True)
    while time.time() < deadline:
        cycle += 1
        if args.max_cycles and cycle > args.max_cycles:
            break
        print(json.dumps({"event": "cycle_start", "cycle": cycle, "remaining_seconds": int(deadline - time.time())}, ensure_ascii=False), flush=True)
        prompts = select_prompts(prompt_scores, cycle)
        cycle_evidence_start = len(evidence)
        deepseek_backend_down_reason = ""

        for target_index, (name, row) in enumerate(targets, start=1):
            if deepseek_backend_down_reason:
                break
            source_path = Path(row["extracted_path"])
            if not source_path.exists():
                append_jsonl(args.output_dir / "failure_corpus.jsonl", {"name": name, "failure_type": "missing_file", "source_name": row["source_name"]})
                continue
            try:
                pages = render_pdf(source_path, rendered_dir)
            except Exception as exc:
                append_jsonl(args.output_dir / "failure_corpus.jsonl", {"name": name, "failure_type": "backend_error", "source_name": row["source_name"], "error": str(exc)})
                continue
            for page in pages:
                try:
                    variants = make_variants(page, variants_dir, cycle)
                except Exception as exc:
                    append_jsonl(args.output_dir / "failure_corpus.jsonl", {"name": name, "failure_type": "backend_error", "source_name": row["source_name"], "error": str(exc)})
                    variants = [("original", page)]

                for variant_name, image_path in variants:
                    if deepseek_backend_down_reason:
                        break
                    for prompt in prompts:
                        if deepseek_backend_down_reason:
                            break
                        started = time.time()
                        text = ""
                        error = ""
                        try:
                            text = post_deepseek(
                                image_path,
                                prompt["text"],
                                args.api_url,
                                args.timeout,
                                max_tokens=args.deepseek_max_tokens,
                                early_stop_account=args.deepseek_early_stop_account,
                                prefix_salvage=args.deepseek_prefix_salvage,
                                repetition_penalty=args.deepseek_repetition_penalty,
                                repetition_context_size=args.deepseek_repetition_context_size,
                            )
                            result = classify_with_labels(text)
                        except Exception as exc:
                            error = f"{type(exc).__name__}: {exc}"
                            result = AccountResult(None, "error", [], error, "deepseek")
                            if is_deepseek_down_error(error):
                                deepseek_backend_down_reason = error
                        elapsed = time.time() - started
                        text_path = text_dir / f"c{cycle:03d}_{target_index:02d}_{safe_filename_part(name)}_deepseek_{variant_name}_{prompt['id']}.txt"
                        text_path.write_text(text, encoding="utf-8")
                        fail = failure_type(result, text, error)
                        if result.value and not looks_like_id_or_timestamp(result.value, text):
                            evidence.append(Evidence(name, result.value, "deepseek", prompt["id"], variant_name, row["source_name"], result.confidence, result.reason, str(text_path)))
                            prompt_scores[prompt["id"]] += 1
                        all_rows.append(
                            {
                                "name": name,
                                "backend": "deepseek",
                                "prompt_id": prompt["id"],
                                "variant": variant_name,
                                "source_name": row["source_name"],
                                "account": result.value or "",
                                "account_masked": mask(result.value or ""),
                                "confidence": result.confidence,
                                "candidates": "; ".join(result.candidates),
                                "candidates_masked": mask("; ".join(result.candidates)),
                                "reason": result.reason,
                                "failure_type": fail,
                                "elapsed_seconds": f"{elapsed:.2f}",
                                "text_path": str(text_path),
                                "image_path": str(image_path),
                                "error": error,
                            }
                        )
                        if fail:
                            append_jsonl(
                                args.output_dir / "failure_corpus.jsonl",
                                {
                                    "name": name,
                                    "backend": "deepseek",
                                    "prompt_id": prompt["id"],
                                    "variant": variant_name,
                                    "source_name": row["source_name"],
                                    "failure_type": fail,
                                    "candidates_masked": mask("; ".join(result.candidates)),
                                    "text_path": str(text_path),
                                    "error": error,
                                },
                            )
                        if deepseek_backend_down_reason:
                            print(
                                json.dumps(
                                    {
                                        "event": "deepseek_backend_down",
                                        "cycle": cycle,
                                        "name": name,
                                        "variant": variant_name,
                                        "prompt_id": prompt["id"],
                                        "error": deepseek_backend_down_reason,
                                    },
                                    ensure_ascii=False,
                                ),
                                flush=True,
                            )

                    if not chandra_disabled_reason:
                        for prompt_id, prompt_text in CHANDRA_PROMPTS:
                            started = time.time()
                            text = ""
                            error = ""
                            try:
                                options = OcrOptions(
                                    backend="chandra",
                                    chandra_method="hf",
                                    chandra_prompt_type="ocr",
                                    chandra_model_checkpoint=args.chandra_model,
                                    chandra_torch_device=args.chandra_device,
                                    chandra_max_output_tokens=args.chandra_max_output_tokens,
                                    chandra_image_max_side=args.chandra_image_max_side,
                                    chandra_prompt=prompt_text,
                                    privacy_check=True,
                                    timeout_seconds=args.timeout,
                                )
                                with time_limit(args.chandra_timeout_seconds, f"Chandra OCR {name} {variant_name}"):
                                    text = extract_text_chandra(image_path, options)
                                result = classify_with_labels(text)
                            except Exception as exc:
                                error = f"{type(exc).__name__}: {exc}"
                                result = AccountResult(None, "error", [], error, "chandra")
                                if (
                                    "Could not initialize Chandra" in error
                                    or "requires chandra" in error
                                    or "OcrTimeoutError" in error
                                ):
                                    chandra_disabled_reason = error
                            elapsed = time.time() - started
                            text_path = text_dir / f"c{cycle:03d}_{target_index:02d}_{safe_filename_part(name)}_chandra_{variant_name}_{prompt_id}.txt"
                            text_path.write_text(text, encoding="utf-8")
                            fail = failure_type(result, text, error)
                            if result.value and not looks_like_id_or_timestamp(result.value, text):
                                evidence.append(Evidence(name, result.value, "chandra", prompt_id, variant_name, row["source_name"], result.confidence, result.reason, str(text_path)))
                            all_rows.append(
                                {
                                    "name": name,
                                    "backend": "chandra",
                                    "prompt_id": prompt_id,
                                    "variant": variant_name,
                                    "source_name": row["source_name"],
                                    "account": result.value or "",
                                    "account_masked": mask(result.value or ""),
                                    "confidence": result.confidence,
                                    "candidates": "; ".join(result.candidates),
                                    "candidates_masked": mask("; ".join(result.candidates)),
                                    "reason": result.reason,
                                    "failure_type": fail,
                                    "elapsed_seconds": f"{elapsed:.2f}",
                                    "text_path": str(text_path),
                                    "image_path": str(image_path),
                                    "error": error,
                                }
                            )
                            if fail:
                                append_jsonl(
                                    args.output_dir / "failure_corpus.jsonl",
                                    {
                                        "name": name,
                                        "backend": "chandra",
                                        "prompt_id": prompt_id,
                                        "variant": variant_name,
                                        "source_name": row["source_name"],
                                        "failure_type": fail,
                                        "candidates_masked": mask("; ".join(result.candidates)),
                                        "text_path": str(text_path),
                                        "error": error,
                                    },
                                )

        chosen = choose_accounts(evidence)
        resolution_path = write_resolution(base_rows, chosen, args.output_dir)
        resolution_rows = read_csv(resolution_path)
        workbook_path = write_workbook(args.source_workbook, resolution_rows, args.output_dir)
        write_csv(args.output_dir / "attempts.csv", all_rows, FIELDS)
        summary = {
            "cycle": cycle,
            "targets": len(targets),
            "missing_files": len(missing_files),
            "new_evidence_this_cycle": len(evidence) - cycle_evidence_start,
            "autoresearch_chosen": len(chosen),
            "best_chosen_count": max(best_chosen_count, len(chosen)),
            "prompt_scores": dict(prompt_scores),
            "chandra_disabled_reason": chandra_disabled_reason,
            "resolution_csv": str(resolution_path),
            "workbook": str(workbook_path),
            "updated_total": sum(1 for row in resolution_rows if row.get("chosen_account")),
            "skipped_total": sum(1 for row in resolution_rows if not row.get("chosen_account")),
            "seconds_left": max(0, int(deadline - time.time())),
        }
        best_chosen_count = max(best_chosen_count, len(chosen))
        (args.output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(json.dumps({"event": "cycle_end", **summary}, ensure_ascii=False), flush=True)

        unresolved_after = [row["name"] for row in resolution_rows if not row.get("chosen_account")]
        if not unresolved_after:
            print(json.dumps({"event": "all_resolved", "cycle": cycle}, ensure_ascii=False), flush=True)
            break
        if time.time() + args.cycle_sleep_seconds >= deadline:
            break
        time.sleep(args.cycle_sleep_seconds)

    print(json.dumps({"event": "finish", "output_dir": str(args.output_dir)}, ensure_ascii=False), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
