#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import mimetypes
import re
import shutil
import subprocess
import sys
import time
import urllib.request
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageEnhance, ImageFilter, ImageOps

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.account_policy import policy_score_candidate  # noqa: E402
from settlement_tool.core import AccountResult, classify_account_candidates, extract_roster, safe_filename_part, write_csv  # noqa: E402
from scripts.build_deepseek_resolution import MANUAL_NAME_HINTS, SOURCE_NAME_HINTS, mask  # noqa: E402


PROMPT_LADDER = [
    {
        "id": "copy_all_text",
        "text": (
            "<image>\n"
            "Copy all visible Korean/English text and numbers from this bankbook/account-copy image. "
            "Preserve line breaks when possible. Do not infer missing text."
        ),
    },
    {
        "id": "account_structured_ko_v2",
        "text": (
            "<image>\n"
            "한국 은행 통장사본 OCR입니다. 보이는 텍스트만 읽으세요. "
            "계좌번호/예금주/은행명을 찾고 추측하지 마세요. "
            "주민등록번호, 운전면허번호, 전화번호, 날짜는 계좌번호가 아닙니다. "
            "출력: bank: ...\\naccount_holder: ...\\naccount_number: ..."
        ),
    },
    {
        "id": "account_only_negative_rules",
        "text": (
            "<image>\n"
            "Find the Korean bank account number only. "
            "Reject phone numbers, dates, timestamps, resident IDs, driver license numbers, and card numbers. "
            "Return exactly one line: account_number: <visible number or UNKNOWN>."
        ),
    },
    {
        "id": "number_inventory_ko",
        "text": (
            "<image>\n"
            "이미지에서 보이는 숫자 묶음을 줄마다 그대로 복사하세요. "
            "각 숫자 옆에 주변 라벨을 함께 적으세요. 예: 계좌번호: 000-000-000000. "
            "설명하지 마세요."
        ),
    },
]


ATTEMPT_FIELDS = [
    "name",
    "backend",
    "prompt_id",
    "variant",
    "source_name",
    "account",
    "account_masked",
    "confidence",
    "candidates",
    "candidates_masked",
    "reason",
    "failure_type",
    "elapsed_seconds",
    "text_path",
    "image_path",
    "error",
]

RESOLUTION_FIELDS = [
    "group",
    "no",
    "name",
    "prior_status",
    "prior_account_masked",
    "chosen_account",
    "chosen_account_masked",
    "decision",
    "source",
    "candidate_count",
    "candidate_accounts_masked",
    "candidate_files",
]


@dataclass(frozen=True)
class RecoveryTarget:
    name: str
    row: dict[str, str]


@dataclass(frozen=True)
class RecoveryManifest:
    active_targets: list[RecoveryTarget]
    missing_files: list[str]
    total_unresolved: int


@dataclass(frozen=True)
class Evidence:
    name: str
    account: str
    prompt_id: str
    variant: str
    source_name: str
    text_path: str


@dataclass(frozen=True)
class ChosenAccount:
    name: str
    account: str
    decision: str
    evidence: list[Evidence]


def read_csv(path: Path) -> list[dict[str, str]]:
    return list(csv.DictReader(path.open(encoding="utf-8-sig")))


def append_jsonl(path: Path, payload: dict[str, object]) -> None:
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")


def run(command: list[str], timeout: int = 120) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, text=True, capture_output=True, check=False, timeout=timeout)


def account_key(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def target_name(row: dict[str, str], target_names: set[str]) -> str:
    source_manual = SOURCE_NAME_HINTS.get(Path(row.get("source_name", "")).stem, "")
    if source_manual in target_names:
        return source_manual
    if row.get("matched_name") in target_names:
        return row["matched_name"]
    manual = MANUAL_NAME_HINTS.get(row.get("filename_hint", ""), "")
    if manual in target_names:
        return manual
    source = row.get("source_name", "")
    for name in target_names:
        if name and name in source:
            return name
    return ""


def build_recovery_manifest(base_rows: list[dict[str, str]], deepseek_rows: list[dict[str, str]]) -> RecoveryManifest:
    unresolved_names = [row["name"] for row in base_rows if not row.get("chosen_account")]
    target_names = set(unresolved_names)
    targets: list[RecoveryTarget] = []
    seen: set[tuple[str, str]] = set()
    for row in deepseek_rows:
        name = target_name(row, target_names)
        if not name:
            continue
        key = (name, row.get("source_name", ""))
        if key in seen:
            continue
        seen.add(key)
        targets.append(RecoveryTarget(name=name, row=row))
    with_files = {target.name for target in targets}
    missing = [name for name in unresolved_names if name not in with_files]
    return RecoveryManifest(active_targets=targets, missing_files=missing, total_unresolved=len(unresolved_names))


def looks_like_id_or_timestamp(candidate: str, text: str) -> bool:
    digits = account_key(candidate)
    if len(digits) == 13:
        marker = digits[6:7]
        if marker in {"1", "2", "3", "4", "5", "6"} and re.fullmatch(r"\d{6}", digits[:6]):
            return True
    if len(digits) >= 10 and digits.startswith("20"):
        return True
    window = text[max(0, text.find(candidate) - 50) : text.find(candidate) + len(candidate) + 50] if candidate in text else text
    id_words = ("주민", "운전", "면허", "등록번호", "license", "resident", "생년월일")
    return any(word.lower() in window.lower() for word in id_words)


def classify_with_labels(text: str) -> AccountResult:
    result = classify_account_candidates(text)
    if result.value:
        return result

    label_re = re.compile(
        r"(?:account[_ ]?number|계좌\s*번호|입금\s*계좌|계좌|Account)\D{0,30}((?:\d[\d -]{7,22}\d))",
        re.IGNORECASE,
    )
    candidates: list[str] = []
    for match in label_re.finditer(text):
        candidate = re.sub(r"[^\d-]", "", match.group(1)).strip("-")
        digits = account_key(candidate)
        decision = policy_score_candidate(text, candidate)
        if 9 <= len(digits) <= 16 and not looks_like_id_or_timestamp(candidate, text) and decision.accepted:
            candidates.append(candidate)
    seen: list[str] = []
    for candidate in candidates:
        if candidate not in seen:
            seen.append(candidate)
    if len(seen) == 1:
        return AccountResult(seen[0], "high", seen, "single_account_label_match", "label_parser")
    if seen:
        return AccountResult(None, "low", seen, "multiple_account_label_matches", "label_parser")
    return result


def failure_type(result: AccountResult, text: str, error: str) -> str:
    if error:
        return "backend_error"
    if result.candidates and any(looks_like_id_or_timestamp(candidate, text) for candidate in result.candidates):
        return "privacy_id_like"
    if result.confidence == "low" and result.candidates:
        return "ambiguous_or_low_confidence"
    if not result.candidates:
        return "ocr_no_candidate"
    return ""


def _ocr_tokens(text: str) -> list[str]:
    return re.findall(r"[A-Za-z가-힣]+|[\u4e00-\u9fff]|\d+", text.lower())


def degenerate_output_reason(text: str) -> str:
    stripped = (text or "").strip()
    if len(stripped) < 12:
        return ""

    digits = account_key(stripped)
    non_space = re.sub(r"\s", "", stripped)
    if len(digits) >= 18 and set(digits) == {"0"}:
        non_zero_noise = re.sub(r"[0.\-:,;()\[\]{}<>/\\|_\s]", "", stripped)
        if len(non_zero_noise) <= max(2, len(non_space) // 20):
            return "zero_placeholder_repetition"

    tokens = _ocr_tokens(stripped)
    if len(tokens) < 8:
        return ""
    unique_ratio = len(set(tokens)) / len(tokens)
    for width in range(1, min(5, len(tokens) + 1)):
        grams = [tuple(tokens[index : index + width]) for index in range(len(tokens) - width + 1)]
        if not grams:
            continue
        phrase, count = Counter(grams).most_common(1)[0]
        coverage = count * width / len(tokens)
        if count >= 4 and coverage >= 0.62 and unique_ratio <= 0.35:
            return "repeated_phrase:" + " ".join(phrase)
    return ""


@dataclass
class DegenerationMonitor:
    max_degenerate_outputs: int
    consecutive_degenerate_outputs: int = 0
    reasons: Counter[str] = field(default_factory=Counter)

    def record(self, text: str) -> str:
        reason = degenerate_output_reason(text)
        if reason:
            self.consecutive_degenerate_outputs += 1
            self.reasons[reason] += 1
        elif text.strip():
            self.consecutive_degenerate_outputs = 0
        return reason

    @property
    def should_stop(self) -> bool:
        return self.max_degenerate_outputs > 0 and self.consecutive_degenerate_outputs >= self.max_degenerate_outputs

    @property
    def reason(self) -> str:
        if not self.should_stop:
            return ""
        top_reason = self.reasons.most_common(1)[0][0] if self.reasons else "unknown"
        return f"{self.consecutive_degenerate_outputs} degenerate DeepSeek outputs in a row: {top_reason}"


def evidence_is_policy_safe(item: Evidence) -> bool:
    path = Path(item.text_path)
    text = path.read_text(encoding="utf-8", errors="replace") if path.exists() else item.account
    return policy_score_candidate(text, item.account).accepted and not looks_like_id_or_timestamp(item.account, text)


def choose_recovery_accounts(evidence: list[Evidence], *, min_consensus: int = 2) -> dict[str, ChosenAccount]:
    by_name_key: dict[tuple[str, str], list[Evidence]] = defaultdict(list)
    for item in evidence:
        if not evidence_is_policy_safe(item):
            continue
        key = account_key(item.account)
        if key:
            by_name_key[(item.name, key)].append(item)

    by_name: dict[str, list[tuple[str, list[Evidence]]]] = defaultdict(list)
    for (name, _), items in by_name_key.items():
        by_name[name].append((items[0].account, items))

    chosen: dict[str, ChosenAccount] = {}
    for name, candidates in by_name.items():
        qualified: list[tuple[str, list[Evidence]]] = []
        for account, items in candidates:
            prompts = {item.prompt_id for item in items}
            variants = {item.variant for item in items}
            if len(items) >= min_consensus and (len(prompts) >= 2 or len(variants) >= 2):
                qualified.append((account, items))
        if len(qualified) == 1 and len(candidates) == 1:
            account, items = qualified[0]
            chosen[name] = ChosenAccount(
                name=name,
                account=account,
                decision="auto_fill_recovery_deepseek_consensus",
                evidence=items,
            )
    return chosen


def render_pdf(path: Path, output_dir: Path) -> list[Path]:
    if path.suffix.lower() != ".pdf":
        return [path]
    if not shutil.which("pdftoppm"):
        raise RuntimeError("pdftoppm is required for PDF OCR")
    prefix = output_dir / safe_filename_part(path.stem)
    proc = run(["pdftoppm", "-png", "-r", "320", str(path), str(prefix)], timeout=180)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or "pdftoppm failed")
    return sorted(output_dir.glob(f"{prefix.name}-*.png"))[:2]


def make_variants(image_path: Path, output_dir: Path) -> list[tuple[str, Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    variants: list[tuple[str, Path]] = [("original", image_path)]
    with Image.open(image_path) as raw:
        img = ImageOps.exif_transpose(raw).convert("RGB")
    width, height = img.size
    if max(width, height) < 1800:
        img = img.resize((width * 2, height * 2), Image.Resampling.LANCZOS)
        width, height = img.size

    gray = ImageOps.grayscale(img)
    gray = ImageOps.autocontrast(gray)
    gray = ImageEnhance.Contrast(gray).enhance(1.8)
    gray = gray.filter(ImageFilter.SHARPEN)
    contrast_path = output_dir / f"{safe_filename_part(image_path.stem)}__contrast.png"
    gray.save(contrast_path)
    variants.append(("contrast", contrast_path))

    top = gray.crop((0, 0, width, int(height * 0.65)))
    if top.size[0] >= 200 and top.size[1] >= 200:
        top_path = output_dir / f"{safe_filename_part(image_path.stem)}__top_65.png"
        top.save(top_path)
        variants.append(("top_65", top_path))
    return variants


def deepseek_form_fields(
    prompt: str,
    *,
    max_tokens: int,
    repetition_penalty: float,
    repetition_context_size: int,
) -> dict[str, str]:
    fields = {
        "content_type": "Scene",
        "subcategory": "Verification",
        "complexity": "Tiny",
        "prompt": prompt,
        "max_tokens": str(max_tokens),
    }
    if repetition_penalty > 0:
        fields["repetition_penalty"] = f"{repetition_penalty:.4g}"
        fields["repetition_context_size"] = str(repetition_context_size)
    return fields


def post_deepseek(
    image_path: Path,
    prompt: str,
    api_url: str,
    timeout: int,
    *,
    max_tokens: int,
    repetition_penalty: float,
    repetition_context_size: int,
) -> str:
    mime = mimetypes.guess_type(image_path.name)[0] or "image/png"
    boundary = f"----recovery-{uuid.uuid4().hex}"
    fields = deepseek_form_fields(
        prompt,
        max_tokens=max_tokens,
        repetition_penalty=repetition_penalty,
        repetition_context_size=repetition_context_size,
    )
    body = bytearray()
    for name, value in fields.items():
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n{value}\r\n'.encode())
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        f'Content-Disposition: form-data; name="file"; filename="{image_path.name}"\r\n'
        f"Content-Type: {mime}\r\n\r\n".encode()
    )
    body.extend(image_path.read_bytes())
    body.extend(b"\r\n")
    body.extend(f"--{boundary}--\r\n".encode())

    request = urllib.request.Request(
        f"{api_url.rstrip('/')}/api/ocr",
        data=bytes(body),
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if not payload.get("success"):
        raise RuntimeError(payload.get("error") or "DeepSeek OCR failed")
    return payload.get("text", "")


def is_deepseek_down_error(error: str) -> bool:
    lowered = error.lower()
    return (
        "connection refused" in lowered
        or "remote end closed connection" in lowered
        or "connection reset" in lowered
        or "broken pipe" in lowered
    )


def write_program(output_dir: Path) -> None:
    (output_dir / "program.md").write_text(
        "\n".join(
            [
                "# Settlement OCR Recovery Loop",
                "",
                "Goal: recover only unresolved Korean bankbook account fields from local DeepSeek MLX OCR.",
                "",
                "Failed hypotheses baked into this loop:",
                "- Missing bankbook files are not OCR-recoverable; record them and skip active OCR.",
                "- Chandra/MPS adds memory pressure; this loop is DeepSeek-only.",
                "- `mode=basic` routes away from Tiny settings; this loop never sends it.",
                "- Blind 12h loops can spam backend errors; stop when the DeepSeek backend is down.",
                "- Degenerate prompt-conditioned text can look like OCR progress; stop after repeated collapsed outputs.",
                "- Single high rows can be prompt leakage or wrong-field extraction; require repeated policy-safe evidence.",
                "- Write target-level checkpoints so panic/reboot does not erase completed attempts.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )


def write_resolution(
    base_rows: list[dict[str, str]],
    chosen: dict[str, ChosenAccount],
    output_dir: Path,
) -> Path:
    rows: list[dict[str, str]] = []
    for row in base_rows:
        current = dict(row)
        picked = chosen.get(current["name"])
        if not current.get("chosen_account") and picked:
            current["chosen_account"] = picked.account
            current["chosen_account_masked"] = mask(picked.account)
            current["decision"] = picked.decision
            current["source"] = ";".join(sorted({f"deepseek:{item.prompt_id}:{item.variant}" for item in picked.evidence}))
            current["candidate_count"] = "1"
            current["candidate_accounts_masked"] = mask(picked.account)
            current["candidate_files"] = "; ".join(sorted({item.source_name for item in picked.evidence}))
        rows.append(current)
    path = output_dir / "account_resolution_candidates.csv"
    write_csv(path, rows, RESOLUTION_FIELDS)
    return path


def write_workbook(source_workbook: Path, resolution_rows: list[dict[str, str]], output_dir: Path) -> Path:
    roster = extract_roster(source_workbook)
    by_name = {row["name"]: row for row in resolution_rows}
    wb = load_workbook(source_workbook, read_only=False, data_only=False)
    ws = wb.active
    report: list[dict[str, str]] = []
    for person in roster.people:
        row = by_name.get(person.name, {})
        account = row.get("chosen_account", "")
        cell = ws[f"J{person.row}"]
        if account:
            cell.value = account
            cell.number_format = "@"
            status = "updated"
        else:
            status = "skipped"
        report.append(
            {
                "group": person.group,
                "no": person.no,
                "name": person.name,
                "cell": cell.coordinate,
                "status": status,
                "decision": row.get("decision", ""),
                "source": row.get("source", ""),
                "account": account,
                "account_masked": row.get("chosen_account_masked", ""),
            }
        )
    path = output_dir / f"{source_workbook.stem}_계좌번호입력_recovery{source_workbook.suffix}"
    wb.save(path)
    if report:
        write_csv(output_dir / "account_updates_recovery.csv", report, list(report[0].keys()))
    return path


def write_missing_files(output_dir: Path, missing_files: list[str]) -> None:
    rows = [{"name": name, "failure_type": "missing_file", "note": "No matching bankbook file in ZIP"} for name in missing_files]
    write_csv(output_dir / "missing_files.csv", rows, ["name", "failure_type", "note"])
    for row in rows:
        append_jsonl(output_dir / "failure_corpus.jsonl", row)


def write_checkpoint(
    *,
    output_dir: Path,
    base_rows: list[dict[str, str]],
    source_workbook: Path,
    attempts: list[dict[str, str]],
    evidence: list[Evidence],
    manifest: RecoveryManifest,
    cycle: int,
    started_at: float,
    backend_down_reason: str = "",
    model_degenerate_reason: str = "",
) -> dict[str, Any]:
    chosen = choose_recovery_accounts(evidence)
    resolution_path = write_resolution(base_rows, chosen, output_dir)
    resolution_rows = read_csv(resolution_path)
    workbook_path = write_workbook(source_workbook, resolution_rows, output_dir)
    write_csv(output_dir / "attempts.csv", attempts, ATTEMPT_FIELDS)
    summary = {
        "cycle": cycle,
        "total_unresolved": manifest.total_unresolved,
        "active_targets": len(manifest.active_targets),
        "missing_files": len(manifest.missing_files),
        "attempts": len(attempts),
        "evidence": len(evidence),
        "recovery_chosen": len(chosen),
        "backend_down_reason": backend_down_reason,
        "model_degenerate_reason": model_degenerate_reason,
        "resolution_csv": str(resolution_path),
        "workbook": str(workbook_path),
        "updated_total": sum(1 for row in resolution_rows if row.get("chosen_account")),
        "skipped_total": sum(1 for row in resolution_rows if not row.get("chosen_account")),
        "elapsed_seconds": round(time.time() - started_at, 2),
        "output_dir": str(output_dir),
    }
    (output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def run_recovery(args: argparse.Namespace) -> int:
    args.output_dir.mkdir(parents=True, exist_ok=True)
    rendered_dir = args.output_dir / "rendered"
    variants_dir = args.output_dir / "variants"
    text_dir = args.output_dir / "ocr_text"
    for path in (rendered_dir, variants_dir, text_dir):
        path.mkdir(parents=True, exist_ok=True)
    write_program(args.output_dir)

    base_rows = read_csv(args.base_resolution_csv)
    deepseek_rows = read_csv(args.deepseek_csv)
    manifest = build_recovery_manifest(base_rows, deepseek_rows)
    if args.limit:
        manifest = RecoveryManifest(
            active_targets=manifest.active_targets[: args.limit],
            missing_files=manifest.missing_files,
            total_unresolved=manifest.total_unresolved,
        )
    write_missing_files(args.output_dir, manifest.missing_files)

    started_at = time.time()
    deadline = started_at + args.duration_hours * 3600
    attempts: list[dict[str, str]] = []
    evidence: list[Evidence] = []
    backend_down_reason = ""
    model_degenerate_reason = ""
    degeneration_monitor = DegenerationMonitor(args.max_degenerate_outputs)

    print(
        json.dumps(
            {
                "event": "start",
                "total_unresolved": manifest.total_unresolved,
                "targets": len(manifest.active_targets),
                "missing_files": manifest.missing_files,
                "output_dir": str(args.output_dir),
            },
            ensure_ascii=False,
        ),
        flush=True,
    )

    cycle = 0
    while time.time() < deadline:
        cycle += 1
        if args.max_cycles and cycle > args.max_cycles:
            break
        print(json.dumps({"event": "cycle_start", "cycle": cycle, "remaining_seconds": int(deadline - time.time())}, ensure_ascii=False), flush=True)

        for target_index, target in enumerate(manifest.active_targets, start=1):
            if backend_down_reason or model_degenerate_reason:
                break
            if target.name in choose_recovery_accounts(evidence):
                continue
            source_path = Path(target.row.get("extracted_path", ""))
            if not source_path.exists():
                append_jsonl(
                    args.output_dir / "failure_corpus.jsonl",
                    {"name": target.name, "failure_type": "missing_file", "source_name": target.row.get("source_name", ""), "note": "Extracted path missing at runtime"},
                )
                continue
            try:
                pages = render_pdf(source_path, rendered_dir)
            except Exception as exc:
                append_jsonl(
                    args.output_dir / "failure_corpus.jsonl",
                    {"name": target.name, "failure_type": "backend_error", "source_name": target.row.get("source_name", ""), "error": f"{type(exc).__name__}: {exc}"},
                )
                continue

            calls_for_target = 0
            for page in pages:
                if backend_down_reason or model_degenerate_reason or calls_for_target >= args.max_calls_per_target:
                    break
                try:
                    variants = make_variants(page, variants_dir)
                except Exception as exc:
                    append_jsonl(
                        args.output_dir / "failure_corpus.jsonl",
                        {"name": target.name, "failure_type": "backend_error", "source_name": target.row.get("source_name", ""), "error": f"{type(exc).__name__}: {exc}"},
                    )
                    variants = [("original", page)]

                for variant_name, image_path in variants:
                    if backend_down_reason or model_degenerate_reason or calls_for_target >= args.max_calls_per_target:
                        break
                    for prompt in PROMPT_LADDER:
                        if backend_down_reason or model_degenerate_reason or calls_for_target >= args.max_calls_per_target:
                            break
                        if target.name in choose_recovery_accounts(evidence, min_consensus=args.min_consensus):
                            break
                        calls_for_target += 1
                        started = time.time()
                        text = ""
                        error = ""
                        try:
                            text = post_deepseek(
                                image_path,
                                prompt["text"],
                                args.api_url,
                                args.timeout,
                                max_tokens=args.deepseek_max_tokens,
                                repetition_penalty=args.deepseek_repetition_penalty,
                                repetition_context_size=args.deepseek_repetition_context_size,
                            )
                            result = classify_with_labels(text)
                        except Exception as exc:
                            error = f"{type(exc).__name__}: {exc}"
                            result = AccountResult(None, "error", [], error, "deepseek")
                            if is_deepseek_down_error(error):
                                backend_down_reason = error
                        elapsed = time.time() - started
                        text_path = text_dir / f"c{cycle:03d}_{target_index:02d}_{safe_filename_part(target.name)}_deepseek_{variant_name}_{prompt['id']}.txt"
                        text_path.write_text(text, encoding="utf-8")
                        degenerate_reason = "" if error else degeneration_monitor.record(text)
                        fail = "model_degenerate_output" if degenerate_reason else failure_type(result, text, error)
                        attempt_reason = result.reason
                        if degenerate_reason:
                            attempt_reason = f"{attempt_reason}; {degenerate_reason}" if attempt_reason else degenerate_reason
                        if result.value and not degenerate_reason and not looks_like_id_or_timestamp(result.value, text):
                            evidence.append(Evidence(target.name, result.value, prompt["id"], variant_name, target.row.get("source_name", ""), str(text_path)))
                        attempts.append(
                            {
                                "name": target.name,
                                "backend": "deepseek",
                                "prompt_id": prompt["id"],
                                "variant": variant_name,
                                "source_name": target.row.get("source_name", ""),
                                "account": result.value or "",
                                "account_masked": mask(result.value or ""),
                                "confidence": result.confidence,
                                "candidates": "; ".join(result.candidates),
                                "candidates_masked": mask("; ".join(result.candidates)),
                                "reason": attempt_reason,
                                "failure_type": fail,
                                "elapsed_seconds": f"{elapsed:.2f}",
                                "text_path": str(text_path),
                                "image_path": str(image_path),
                                "error": error,
                            }
                        )
                        if fail:
                            append_jsonl(
                                args.output_dir / "failure_corpus.jsonl",
                                {
                                    "name": target.name,
                                    "backend": "deepseek",
                                    "prompt_id": prompt["id"],
                                    "variant": variant_name,
                                    "source_name": target.row.get("source_name", ""),
                                    "failure_type": fail,
                                    "degenerate_reason": degenerate_reason,
                                    "candidates_masked": mask("; ".join(result.candidates)),
                                    "text_path": str(text_path),
                                    "error": error,
                                },
                            )
                        if degeneration_monitor.should_stop and not model_degenerate_reason:
                            model_degenerate_reason = degeneration_monitor.reason
                        summary = write_checkpoint(
                            output_dir=args.output_dir,
                            base_rows=base_rows,
                            source_workbook=args.source_workbook,
                            attempts=attempts,
                            evidence=evidence,
                            manifest=manifest,
                            cycle=cycle,
                            started_at=started_at,
                            backend_down_reason=backend_down_reason,
                            model_degenerate_reason=model_degenerate_reason,
                        )
                        print(json.dumps({"event": "attempt", "cycle": cycle, "name": target.name, "prompt_id": prompt["id"], "variant": variant_name, "recovery_chosen": summary["recovery_chosen"]}, ensure_ascii=False), flush=True)
                        if model_degenerate_reason:
                            print(json.dumps({"event": "deepseek_model_degenerate", "reason": model_degenerate_reason}, ensure_ascii=False), flush=True)

        summary = write_checkpoint(
            output_dir=args.output_dir,
            base_rows=base_rows,
            source_workbook=args.source_workbook,
            attempts=attempts,
            evidence=evidence,
            manifest=manifest,
            cycle=cycle,
            started_at=started_at,
            backend_down_reason=backend_down_reason,
            model_degenerate_reason=model_degenerate_reason,
        )
        print(json.dumps({"event": "cycle_end", **summary}, ensure_ascii=False), flush=True)
        if backend_down_reason:
            print(json.dumps({"event": "deepseek_backend_down", "error": backend_down_reason}, ensure_ascii=False), flush=True)
            break
        if model_degenerate_reason:
            break
        if time.time() + args.cycle_sleep_seconds >= deadline:
            break
        time.sleep(args.cycle_sleep_seconds)

    print(json.dumps({"event": "finish", "output_dir": str(args.output_dir)}, ensure_ascii=False), flush=True)
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run DeepSeek-only account recovery for unresolved settlement OCR rows.")
    parser.add_argument("--source-workbook", type=Path, required=True)
    parser.add_argument("--base-resolution-csv", type=Path, required=True)
    parser.add_argument("--deepseek-csv", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--duration-hours", type=float, default=12)
    parser.add_argument("--api-url", default="http://127.0.0.1:5001")
    parser.add_argument("--timeout", type=int, default=900)
    parser.add_argument("--cycle-sleep-seconds", type=int, default=60)
    parser.add_argument("--deepseek-max-tokens", type=int, default=256)
    parser.add_argument("--deepseek-repetition-penalty", type=float, default=1.05)
    parser.add_argument("--deepseek-repetition-context-size", type=int, default=64)
    parser.add_argument("--max-cycles", type=int, default=1)
    parser.add_argument("--max-calls-per-target", type=int, default=9)
    parser.add_argument("--max-degenerate-outputs", type=int, default=6)
    parser.add_argument("--min-consensus", type=int, default=2)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--disable-chandra", action="store_true", help=argparse.SUPPRESS)
    return parser.parse_args()


def main() -> int:
    return run_recovery(parse_args())


if __name__ == "__main__":
    raise SystemExit(main())
