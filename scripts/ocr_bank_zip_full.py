#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from settlement_tool.core import (  # noqa: E402
    AccountResult,
    classify_account_candidates,
    decode_zip_name,
    extract_roster,
    normalize_text,
    safe_filename_part,
    write_csv,
)


@dataclass(frozen=True)
class ZipMember:
    source_name: str
    size: int
    date_time: tuple[int, int, int, int, int, int]
    suffix: str


def read_form_rows(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = normalize_text(row[1] if len(row) > 1 else "")
        if not name or name == "테스트":
            continue
        rows.append(
            {
                "form_row": str(index),
                "timestamp": normalize_text(row[0] if len(row) > 0 else ""),
                "name": name,
                "phone": normalize_text(row[2] if len(row) > 2 else ""),
                "role": normalize_text(row[3] if len(row) > 3 else ""),
                "id_link": normalize_text(row[5] if len(row) > 5 else ""),
                "bank_link": normalize_text(row[6] if len(row) > 6 else ""),
            }
        )
    return rows


def zip_members(path: Path) -> list[ZipMember]:
    members = []
    with zipfile.ZipFile(path) as zf:
        for info in zf.infolist():
            decoded = decode_zip_name(info.filename)
            if info.is_dir() or Path(decoded).name == ".DS_Store":
                continue
            members.append(
                ZipMember(
                    source_name=decoded,
                    size=info.file_size,
                    date_time=info.date_time,
                    suffix=Path(decoded).suffix.lower(),
                )
            )
    return members


def extract_member(zip_path: Path, member: ZipMember, destination: Path) -> Path:
    target = normalize_text(member.source_name)
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if decode_zip_name(info.filename) == target:
                destination.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(info) as source, destination.open("wb") as dest:
                    shutil.copyfileobj(source, dest)
                return destination
    raise FileNotFoundError(member.source_name)


def run(command: list[str], timeout: int = 120) -> subprocess.CompletedProcess[str]:
    return subprocess.run(command, text=True, capture_output=True, check=False, timeout=timeout)


def ocr_file(path: Path, work_dir: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text = ""
        if shutil.which("pdftotext"):
            proc = run(["pdftotext", "-layout", str(path), "-"], timeout=60)
            if proc.returncode == 0:
                text = proc.stdout
        if text.strip():
            return text
        if not shutil.which("pdftoppm"):
            return ""
        prefix = work_dir / f"{path.stem}_page"
        proc = run(["pdftoppm", "-png", "-r", "220", str(path), str(prefix)], timeout=120)
        if proc.returncode != 0:
            return ""
        return "\n".join(ocr_image(image) for image in sorted(work_dir.glob(f"{path.stem}_page-*.png")))
    return ocr_image(path)


def ocr_image(path: Path) -> str:
    if not shutil.which("tesseract"):
        return ""
    outputs = []
    for psm in ("6", "11"):
        proc = run(["tesseract", str(path), "stdout", "-l", "kor+eng", "--psm", psm], timeout=90)
        if proc.returncode == 0 and proc.stdout.strip():
            outputs.append(proc.stdout)
    return "\n".join(outputs)


def compact(value: str) -> str:
    return re.sub(r"[\s_()\\[\\]\\-]+", "", normalize_text(value)).lower()


def basename_hint(source_name: str) -> str:
    stem = Path(source_name).stem
    if " - " in stem:
        return normalize_text(stem.rsplit(" - ", 1)[-1])
    return normalize_text(stem)


def score_name(name: str, member: ZipMember, ocr_text: str) -> int:
    c_name = compact(name)
    source = compact(Path(member.source_name).name)
    hint = compact(basename_hint(member.source_name))
    ocr = compact(ocr_text)
    score = 0
    if c_name and c_name in source:
        score += 80
    if c_name and c_name in hint:
        score += 100
    if c_name and c_name in ocr:
        score += 80
    return score


def best_name_match(names: list[str], member: ZipMember, ocr_text: str) -> tuple[str, int, str]:
    scored = sorted(
        ((name, score_name(name, member, ocr_text)) for name in names),
        key=lambda item: item[1],
        reverse=True,
    )
    if not scored or scored[0][1] <= 0:
        return "", 0, "no_name_signal"
    if len(scored) > 1 and scored[1][1] == scored[0][1]:
        return scored[0][0], scored[0][1], "ambiguous_name_score"
    reason = "filename_or_ocr_name_match"
    return scored[0][0], scored[0][1], reason


def candidate_summary(result: AccountResult) -> str:
    return "; ".join(result.candidates)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--form-xlsx", type=Path, required=True)
    parser.add_argument("--bank-zip", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    extracted_dir = args.output_dir / "bank_zip_extracted"
    ocr_text_dir = args.output_dir / "ocr_text"
    extracted_dir.mkdir(parents=True, exist_ok=True)
    ocr_text_dir.mkdir(parents=True, exist_ok=True)

    roster = extract_roster(args.workbook)
    roster_names = roster.names
    form_rows = read_form_rows(args.form_xlsx)
    form_names = [row["name"] for row in form_rows]
    name_pool = sorted(set(roster_names) | set(form_names))

    rows = []
    with tempfile.TemporaryDirectory(prefix="bank_zip_full_ocr_") as tmp:
        tmp_dir = Path(tmp)
        members = zip_members(args.bank_zip)
        for index, member in enumerate(members, start=1):
            print(f"OCR {index}/{len(members)} {Path(member.source_name).name}", flush=True)
            safe_stem = f"{index:03d}_{safe_filename_part(Path(member.source_name).stem)}"
            local_path = extracted_dir / f"{safe_stem}{Path(member.source_name).suffix}"
            extract_member(args.bank_zip, member, local_path)
            text = ocr_file(local_path, tmp_dir)
            (ocr_text_dir / f"{safe_stem}.txt").write_text(text, encoding="utf-8")

            account_result = classify_account_candidates(text)
            matched_name, name_score, match_reason = best_name_match(name_pool, member, text)
            form_match = next((row for row in form_rows if row["name"] == matched_name), None)
            roster_match = next((person for person in roster.people if person.name == matched_name), None)
            dt = member.date_time
            rows.append(
                {
                    "index": index,
                    "zip_datetime": f"{dt[0]:04d}-{dt[1]:02d}-{dt[2]:02d} {dt[3]:02d}:{dt[4]:02d}:{dt[5]:02d}",
                    "source_name": member.source_name,
                    "extracted_path": str(local_path),
                    "filename_hint": basename_hint(member.source_name),
                    "matched_name": matched_name,
                    "name_score": name_score,
                    "match_reason": match_reason,
                    "form_row": form_match["form_row"] if form_match else "",
                    "form_role": form_match["role"] if form_match else "",
                    "roster_group": roster_match.group if roster_match else "",
                    "roster_no": roster_match.no if roster_match else "",
                    "account": account_result.value or "",
                    "confidence": account_result.confidence,
                    "candidates": candidate_summary(account_result),
                    "account_reason": account_result.reason,
                    "ocr_text_path": str(ocr_text_dir / f"{safe_stem}.txt"),
                }
            )

    write_csv(
        args.output_dir / "bank_zip_full_ocr.csv",
        rows,
        [
            "index",
            "zip_datetime",
            "source_name",
            "extracted_path",
            "filename_hint",
            "matched_name",
            "name_score",
            "match_reason",
            "form_row",
            "form_role",
            "roster_group",
            "roster_no",
            "account",
            "confidence",
            "candidates",
            "account_reason",
            "ocr_text_path",
        ],
    )

    summary = {
        "members": len(rows),
        "high_accounts": sum(1 for row in rows if row["confidence"] == "high"),
        "matched_names": sum(1 for row in rows if row["matched_name"]),
        "output": str(args.output_dir),
    }
    (args.output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
