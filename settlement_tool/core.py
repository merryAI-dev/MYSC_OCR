from __future__ import annotations

import csv
import json
import re
import shutil
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .account_policy import policy_score_candidate, rerank_account_candidates


DOC_PAYMENT = "비용지급확인서"
DOC_ID = "신분증"
DOC_BANK = "통장사본"
DOC_TYPES = (DOC_PAYMENT, DOC_ID, DOC_BANK)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return unicodedata.normalize("NFC", str(value)).strip()


def compact_text(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def safe_filename_part(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r'[\\/:*?"<>|]', "_", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value or "unknown"


@dataclass(frozen=True)
class Person:
    group: str
    no: int
    name: str
    row: int


@dataclass
class Roster:
    people: list[Person]

    @property
    def names(self) -> list[str]:
        return [person.name for person in self.people]

    def by_name(self) -> dict[str, Person]:
        return {person.name: person for person in self.people}

    def group_counts(self) -> dict[str, int]:
        return dict(Counter(person.group for person in self.people))


@dataclass(frozen=True)
class FileRef:
    source_name: str
    path: Path | None = None
    zip_path: Path | None = None

    @property
    def suffix(self) -> str:
        return Path(self.source_name).suffix


@dataclass
class MatchResult:
    confirmed: dict[str, FileRef] = field(default_factory=dict)
    ambiguous: dict[str, list[FileRef]] = field(default_factory=dict)
    missing: list[str] = field(default_factory=list)
    unmatched_files: list[str] = field(default_factory=list)


@dataclass
class AccountResult:
    value: str | None
    confidence: str
    candidates: list[str]
    reason: str
    backend: str = ""


def extract_roster(workbook_path: Path | str) -> Roster:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=False, data_only=False)
    ws = wb.active
    people: list[Person] = []
    current_group = ""

    for row in ws.iter_rows(values_only=False):
        b_value = normalize_text(row[1].value) if len(row) > 1 else ""
        name = normalize_text(row[3].value) if len(row) > 3 else ""

        if re.match(r"^\d+\.\s*", b_value) and "인)" in b_value:
            current_group = b_value
            continue

        no_value = row[1].value if len(row) > 1 else None
        if isinstance(no_value, (int, float)) and name and name != "성명":
            people.append(
                Person(
                    group=current_group,
                    no=int(no_value),
                    name=name,
                    row=row[0].row,
                )
            )

    return Roster(people=people)


def decode_zip_name(filename: str) -> str:
    try:
        return normalize_text(filename.encode("cp437").decode("utf-8"))
    except (UnicodeEncodeError, UnicodeDecodeError):
        return normalize_text(filename)


def zip_file_names(zip_path: Path | str) -> list[str]:
    zip_path = Path(zip_path)
    with zipfile.ZipFile(zip_path) as zf:
        return [
            decode_zip_name(info.filename)
            for info in zf.infolist()
            if not info.is_dir() and Path(decode_zip_name(info.filename)).name != ".DS_Store"
        ]


def match_files_by_name(names: Iterable[str], file_names: Iterable[str]) -> MatchResult:
    normalized_names = [normalize_text(name) for name in names]
    files = [normalize_text(file_name) for file_name in file_names]
    result = MatchResult()
    used_files: set[str] = set()

    for name in normalized_names:
        needle = compact_text(name)
        candidates = [
            FileRef(source_name=file_name)
            for file_name in files
            if needle and needle in compact_text(file_name)
        ]
        if len(candidates) == 1:
            result.confirmed[name] = candidates[0]
            used_files.add(candidates[0].source_name)
        elif len(candidates) > 1:
            result.ambiguous[name] = candidates
            used_files.update(candidate.source_name for candidate in candidates)
        else:
            result.missing.append(name)

    result.unmatched_files = [file_name for file_name in files if file_name not in used_files]
    return result


def _normalize_account_candidate(value: str) -> str:
    value = re.sub(r"[^\d-]", "", value)
    value = re.sub(r"-{2,}", "-", value).strip("-")
    return value


def _looks_like_phone_number(candidate: str) -> bool:
    parts = candidate.split("-")
    digits = candidate.replace("-", "")
    if re.fullmatch(r"010-?\d{4}-?\d{4}", candidate):
        return True
    if re.fullmatch(r"0(?:2|[3-6][1-5]|70|80)-?\d{3,4}-?\d{4}", candidate):
        return True
    if re.fullmatch(r"82-?\d{1,2}-?\d{3,4}-?\d{4}", candidate):
        return True
    if len(parts) == 3 and parts[0] in {"02", "070", "080"}:
        return True
    if len(parts) == 3 and re.fullmatch(r"0[3-6][1-5]", parts[0]):
        return True
    if len(digits) in {9, 10, 11} and digits.startswith(("02", "070", "080", "050", "031", "032", "033", "041", "042", "043", "044", "051", "052", "053", "054", "055", "061", "062", "063", "064")):
        return True
    return False


def _account_candidate_score(text: str, candidate: str) -> int:
    return int(policy_score_candidate(text, candidate).score)


def classify_account_candidates(text: str) -> AccountResult:
    text = normalize_text(text)
    raw_candidates = re.findall(r"(?<!\d)(?:\d[\d -]{7,22}\d)(?!\d)", text)
    seen: set[str] = set()
    candidates: list[str] = []

    for raw in raw_candidates:
        candidate = _normalize_account_candidate(raw)
        digits = candidate.replace("-", "")
        if candidate in seen:
            continue
        seen.add(candidate)
        if len(digits) < 9 or len(digits) > 16:
            continue
        if _looks_like_phone_number(candidate):
            continue
        if re.fullmatch(r"20\d{6,12}", digits):
            continue
        candidates.append(candidate)

    if not candidates:
        return AccountResult(None, "none", [], "no_account_candidate")

    reranked = rerank_account_candidates(text, candidates)
    ranked_candidates = [decision.candidate for decision in reranked.decisions]
    if reranked.selected is not None:
        return AccountResult(reranked.selected.candidate, "high", ranked_candidates, reranked.reason)

    return AccountResult(None, "low", ranked_candidates, reranked.reason)


def parse_overrides(path: Path | str | None) -> dict[tuple[str, str], str]:
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    overrides: dict[tuple[str, str], str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = normalize_text(row.get("name"))
            field = normalize_text(row.get("field"))
            value = normalize_text(row.get("value"))
            if name and field and value:
                overrides[(name, field)] = value
    return overrides


def load_config(path: Path | str) -> dict[str, object]:
    with Path(path).open("r", encoding="utf-8") as handle:
        return json.load(handle)


def timestamped_output_dir(base_dir: Path | str) -> Path:
    return Path(base_dir) / datetime.now().strftime("%Y%m%d_%H%M%S")


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def extract_zip_member(zip_path: Path, member_name: str, destination: Path) -> None:
    decoded_target = normalize_text(member_name)
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if decode_zip_name(info.filename) == decoded_target:
                ensure_dir(destination.parent)
                with zf.open(info) as source, destination.open("wb") as dest:
                    shutil.copyfileobj(source, dest)
                return
    raise FileNotFoundError(f"{member_name} not found in {zip_path}")


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def write_text(path: Path, text: str) -> None:
    ensure_dir(path.parent)
    path.write_text(text, encoding="utf-8")


def doc_filename(name: str, doc_type: str, suffix: str) -> str:
    suffix = suffix if suffix.startswith(".") else f".{suffix}" if suffix else ""
    return f"{safe_filename_part(name)}_{doc_type}{suffix}"
