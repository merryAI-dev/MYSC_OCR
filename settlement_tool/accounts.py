from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .core import AccountResult, extract_roster, normalize_text, write_csv


def apply_account_updates(
    source_workbook: Path | str,
    output_workbook: Path | str,
    account_results: dict[str, AccountResult],
) -> list[dict[str, object]]:
    source_workbook = Path(source_workbook)
    output_workbook = Path(output_workbook)
    roster = extract_roster(source_workbook)
    wb = load_workbook(source_workbook, read_only=False, data_only=False)
    ws = wb.active
    report: list[dict[str, object]] = []

    for person in roster.people:
        result = account_results.get(person.name)
        cell = ws[f"J{person.row}"]
        if result and result.value and result.confidence == "high":
            cell.value = result.value
            cell.number_format = "@"
            status = "updated"
        else:
            status = "skipped"
        report.append(
            {
                "group": person.group,
                "no": person.no,
                "name": person.name,
                "cell": cell.coordinate,
                "status": status,
                "account": normalize_text(result.value) if result and result.value else "",
                "confidence": result.confidence if result else "none",
                "candidates": "; ".join(result.candidates) if result else "",
                "reason": result.reason if result else "no_result",
                "backend": result.backend if result else "",
            }
        )

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_workbook)
    return report


def write_account_report(report: list[dict[str, object]], path: Path | str) -> None:
    write_csv(
        Path(path),
        report,
        ["group", "no", "name", "cell", "status", "account", "confidence", "candidates", "reason", "backend"],
    )
