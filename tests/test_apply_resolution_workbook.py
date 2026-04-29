import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.apply_resolution_workbook import apply_resolution_workbook, is_workbook_auto_fill_decision


def test_workbook_auto_fill_decisions_include_policy_reranker_outputs():
    assert is_workbook_auto_fill_decision("keep_existing_final_run") is True
    assert is_workbook_auto_fill_decision("auto_fill_single_candidate") is True
    assert is_workbook_auto_fill_decision("auto_fill_targeted_deepseek") is True
    assert is_workbook_auto_fill_decision("auto_fill_policy_reranker") is True
    assert is_workbook_auto_fill_decision("auto_fill_targeted_policy_reranker") is True
    assert is_workbook_auto_fill_decision("auto_fill_openai_reranker") is True
    assert is_workbook_auto_fill_decision("multiple_candidates_review") is False


def _workbook_with_review_sheet(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "시트"
    ws["B3"] = "1. 테스트 (2인)"
    ws["B5"] = "No."
    ws["D5"] = "성명"
    ws["J5"] = "계좌번호"
    for row_index, no, name in ((6, 1, "홍길동"), (7, 2, "최검토")):
        ws.cell(row_index, 2).value = no
        ws.cell(row_index, 4).value = name

    review = wb.create_sheet("REMAINING_REVIEW")
    review.append(["group", "no", "name", "reason"])
    review.append(["1. 테스트 (2인)", 2, "최검토", "holder_mismatch"])
    wb.save(path)


def test_apply_resolution_workbook_enforces_manual_review_hard_gate(tmp_path: Path):
    source_workbook = tmp_path / "source.xlsx"
    resolution_csv = tmp_path / "resolution.csv"
    output_dir = tmp_path / "out"
    _workbook_with_review_sheet(source_workbook)
    resolution_csv.write_text(
        "name,decision,chosen_account,chosen_account_masked,source,candidate_count,candidate_accounts_masked,candidate_files\n"
        "홍길동,auto_fill_openai_reranker,110-123-456789,***-***-**6789,openai,1,***-***-**6789,a.png\n"
        "최검토,auto_fill_openai_reranker,222-333-444444,***-***-**4444,openai,1,***-***-**4444,b.png\n",
        encoding="utf-8",
    )

    summary = apply_resolution_workbook(
        source_workbook=source_workbook,
        resolution_csv=resolution_csv,
        output_dir=output_dir,
        manual_review_workbook=source_workbook,
    )
    rows = list(csv.DictReader((output_dir / "account_updates_deepseek_resolution.csv").open(encoding="utf-8-sig")))
    wb = load_workbook(summary["output_workbook"], data_only=False)
    ws = wb.active

    assert summary["updated"] == 1
    assert summary["hard_gate_blocked"] == 1
    assert ws["J6"].value == "110-123-456789"
    assert ws["J7"].value is None
    assert rows[0]["status"] == "updated"
    assert rows[1]["status"] == "hard_gate_blocked"
    assert rows[1]["hard_gate_reason"] == "manual_review_required"
