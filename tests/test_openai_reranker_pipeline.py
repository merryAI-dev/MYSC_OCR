import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook

from scripts.apply_openai_reranker_decisions import build_openai_resolution_candidates
from scripts.build_openai_reranker_payloads import build_openai_reranker_files, build_openai_reranker_payloads
from scripts.compare_human_eval_reports import compare_human_eval_reports
from scripts.local_cross_encoder_rerank import candidate_document, rerank_payload_with_scores
from scripts.local_feature_score_rerank import rerank_payload_with_v2_feature_score, v2_feature_score
from scripts.local_oss_structured_rerank import _extract_json_object, _repair_decision
from scripts.openai_structured_rerank import dry_run_rerank_payload, validate_decision
from settlement_tool.teacher_distill import build_candidate_features


RAW_ACCOUNT_RE = re.compile(r"(?<![0-9.])(?:[0-9][0-9 -]{7,22}[0-9])(?![0-9.])")


def test_build_openai_reranker_payloads_keep_raw_pii_in_local_map_only(tmp_path: Path):
    [feature] = build_candidate_features(
        "예금주 홍길동\n연락처 010-1234-5678\n국민은행 계좌번호 110-123-456789",
        source_id="source-1",
        source_name="/Users/boram/Downloads/홍길동_통장.png",
        backend="mlx-deepseek",
        variant="contrast",
        prompt_id="account_structured_ko",
    )
    feature["matched_name"] = "홍길동"

    payloads, raw_maps = build_openai_reranker_payloads([feature])
    payload_text = json.dumps(payloads, ensure_ascii=False)
    raw_map_text = json.dumps(raw_maps, ensure_ascii=False)

    assert payloads[0]["source_id"] == "source-1"
    assert payloads[0]["candidates"][0]["candidate_id"] == "acct_1"
    assert payloads[0]["candidates"][0]["account_shape"] == "ddd-ddd-dddddd"
    assert not RAW_ACCOUNT_RE.search(payload_text)
    assert "110-123-456789" not in payload_text
    assert "010-1234-5678" not in payload_text
    assert "홍길동" not in payload_text
    assert "/Users/boram" not in payload_text
    assert "110-123-456789" in raw_map_text
    assert "홍길동" in raw_map_text


def test_build_openai_reranker_payloads_splits_same_source_by_person_without_leaking_names():
    [first] = build_candidate_features(
        "국민은행 계좌번호 110-123-456789",
        source_id="source-1",
        source_name="/Users/boram/Downloads/shared.png",
    )
    [second] = build_candidate_features(
        "하나은행 계좌번호 222-333-444444",
        source_id="source-1",
        source_name="/Users/boram/Downloads/shared.png",
    )
    first["matched_name"] = "홍길동"
    second["matched_name"] = "김철수"

    payloads, raw_maps = build_openai_reranker_payloads([first, second])
    payload_text = json.dumps(payloads, ensure_ascii=False)
    payload_source_ids = {payload["source_id"] for payload in payloads}

    assert len(payloads) == 2
    assert [len(payload["candidates"]) for payload in payloads] == [1, 1]
    assert {row["source_id"] for row in raw_maps} == payload_source_ids
    assert "홍길동" not in payload_text
    assert "김철수" not in payload_text
    assert "/Users/boram" not in payload_text
    assert all("홍길동" not in source_id and "김철수" not in source_id for source_id in payload_source_ids)


def test_build_openai_reranker_files_writes_redacted_payload_and_local_map(tmp_path: Path):
    features_path = tmp_path / "candidate_features_local.jsonl"
    feature = build_candidate_features(
        "국민은행 계좌번호 110-123-456789",
        source_id="source-1",
        source_name="홍길동_통장.png",
    )[0]
    feature["matched_name"] = "홍길동"
    features_path.write_text(json.dumps(feature, ensure_ascii=False) + "\n", encoding="utf-8")

    summary = build_openai_reranker_files(features_jsonl=features_path, output_dir=tmp_path / "openai")
    payload_text = Path(summary["outputs"]["candidate_features_redacted"]).read_text(encoding="utf-8")
    local_map_text = Path(summary["outputs"]["candidate_raw_map_local"]).read_text(encoding="utf-8")

    assert summary["payload_count"] == 1
    assert "110-123-456789" not in payload_text
    assert "홍길동" not in payload_text
    assert "110-123-456789" in local_map_text


def test_v3_redacted_payload_adds_shape_field_holder_and_consensus_features_without_pii():
    full_feature = build_candidate_features(
        "예금주 홍길동\n국민은행\n계좌번호 110-123-456789",
        source_id="deepseek_bank_zip_full_ocr:1",
        source_name="/Users/boram/Downloads/홍길동_통장.png",
        variant="original",
        prompt_id="account_structured_ko",
    )[0]
    retry_feature = build_candidate_features(
        "<tr><td>예금주</td><td>홍길동</td></tr><tr><td>은행명</td><td>국민은행</td></tr><tr><td>계좌번호</td><td>110-123-456789</td></tr>",
        source_id="targeted_retry_ocr:2",
        source_name="/Users/boram/Downloads/홍길동_통장.png",
        variant="contrast",
        prompt_id="account_structured_ko",
    )[0]
    for feature in (full_feature, retry_feature):
        feature["matched_name"] = "홍길동"

    payloads, _ = build_openai_reranker_payloads([full_feature, retry_feature])
    payload_text = json.dumps(payloads, ensure_ascii=False)
    candidate = payloads[0]["candidates"][0]

    assert payloads[0]["schema_version"] == "openai_account_reranker_redacted_v3"
    assert candidate["shape_features"]["group_lengths"] == [3, 3, 6]
    assert candidate["shape_features"]["digit_count_bucket"] == "11_13"
    assert candidate["shape_features"]["pattern_family"] == "bank_account_like"
    assert candidate["field_evidence"]["same_line_label_type"] == "account_number"
    assert candidate["field_evidence"]["is_value_in_account_field"] is True
    assert candidate["bank_holder_evidence"]["bank_name_present"] is True
    assert candidate["bank_holder_evidence"]["bank_name_normalized"] == "KB국민은행"
    assert candidate["bank_holder_evidence"]["holder_field_present"] is True
    assert candidate["bank_holder_evidence"]["holder_match_status"] == "match"
    assert candidate["consensus_features"]["seen_in_full_ocr"] is True
    assert candidate["consensus_features"]["seen_in_targeted_retry"] is True
    assert candidate["consensus_features"]["variant_vote_count"] == 2
    assert candidate["consensus_features"]["same_candidate_seen_across_variants"] is True
    assert candidate["source_evidence"]["source_kind"] == "full_ocr"
    assert candidate["source_evidence"]["variant"] == "original"
    assert candidate["source_evidence"]["prompt_id"] == "account_structured_ko"
    assert "110-123-456789" not in payload_text
    assert "홍길동" not in payload_text
    assert "/Users/boram" not in payload_text


def test_v3_payload_includes_kie_layout_evidence_without_raw_values():
    feature = {
        "source_id": "kie:1",
        "source_name": "/Users/boram/private/source.png",
        "candidate_raw": "RAW_ACCOUNT_SENTINEL",
        "candidate_masked": "***-***-**6789",
        "digit_count": 12,
        "hyphen_count": 2,
        "group_count": 3,
        "teacher_policy_score": 14,
        "repeat_count": 1,
        "kie_backend": "paddleocr_kie",
        "kie_field_type": "account_number",
        "kie_label_masked": "계좌번호",
        "kie_confidence": 0.93,
        "kie_confidence_bucket": "high",
        "kie_holder_match_status": "match",
        "kie_holder_field_present": True,
        "kie_bank_name_present": True,
        "layout_evidence": {"x_bucket": "left", "y_bucket": "top"},
        "matched_name": "LOCAL_NAME_SENTINEL",
    }

    payloads, raw_maps = build_openai_reranker_payloads([feature])
    text = json.dumps(payloads, ensure_ascii=False)
    candidate = payloads[0]["candidates"][0]

    assert payloads[0]["schema_version"] == "openai_account_reranker_redacted_v3"
    assert candidate["kie_evidence"]["field_type"] == "account_number"
    assert candidate["kie_evidence"]["backend"] == "paddleocr_kie"
    assert candidate["bank_holder_evidence"]["holder_match_status"] == "match"
    assert candidate["bank_holder_evidence"]["bank_name_present"] is True
    assert candidate["layout_evidence"]["x_bucket"] == "left"
    assert "RAW_ACCOUNT_SENTINEL" not in text
    assert "LOCAL_NAME_SENTINEL" not in text
    assert "/Users/boram" not in text
    assert raw_maps[0]["candidate_raw"] == "RAW_ACCOUNT_SENTINEL"


def test_v3_payload_does_not_mark_regular_ocr_features_as_kie_evidence():
    [feature] = build_candidate_features(
        "국민은행 계좌번호 110-123-456789",
        source_id="source-1",
        source_name="sample.png",
        backend="mixed_candidate_generation",
    )

    payloads, _ = build_openai_reranker_payloads([feature])
    candidate = payloads[0]["candidates"][0]

    assert candidate["kie_evidence"]["backend"] == ""
    assert candidate["kie_evidence"]["field_type"] == "unknown"


def test_dry_run_openai_structured_reranker_accepts_only_margin_winner():
    payload = {
        "source_id": "source-1",
        "candidates": [
            {
                "candidate_id": "acct_1",
                "teacher_policy_score": 18.0,
                "risk_flags": {
                    "looks_like_phone": False,
                    "has_prompt_leakage_context": False,
                    "has_wrong_field_context": False,
                },
            },
            {
                "candidate_id": "acct_2",
                "teacher_policy_score": 13.0,
                "risk_flags": {
                    "looks_like_phone": False,
                    "has_prompt_leakage_context": False,
                    "has_wrong_field_context": False,
                },
            },
        ],
    }

    decision = dry_run_rerank_payload(payload, threshold=10.0, min_margin=3.0)

    assert decision["action"] == "accept"
    assert decision["selected_candidate_id"] == "acct_1"
    validate_decision(payload, decision)


def test_apply_openai_reranker_decisions_maps_raw_locally_and_enforces_hard_gate(tmp_path: Path):
    workbook = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "시트"
    ws["B3"] = "1. 테스트 (2인)"
    ws["B5"] = "No."
    ws["D5"] = "성명"
    ws.cell(6, 2).value = 1
    ws.cell(6, 4).value = "홍길동"
    ws.cell(7, 2).value = 2
    ws.cell(7, 4).value = "최검토"
    review = wb.create_sheet("REMAINING_REVIEW")
    review.append(["group", "no", "name", "reason"])
    review.append(["1. 테스트 (2인)", 2, "최검토", "holder_mismatch"])
    wb.save(workbook)

    raw_map = tmp_path / "candidate_raw_map_local.jsonl"
    raw_map.write_text(
        json.dumps(
            {
                "source_id": "source-1",
                "candidate_id": "acct_1",
                "name": "홍길동",
                "candidate_raw": "110-123-456789",
                "candidate_masked": "***-***-**6789",
                "source_name": "a.png",
            },
            ensure_ascii=False,
        )
        + "\n"
        + json.dumps(
            {
                "source_id": "source-2",
                "candidate_id": "acct_1",
                "name": "최검토",
                "candidate_raw": "222-333-444444",
                "candidate_masked": "***-***-**4444",
                "source_name": "b.png",
            },
            ensure_ascii=False,
        )
        + "\n",
        encoding="utf-8",
    )
    decisions = tmp_path / "decisions.jsonl"
    decisions.write_text(
        json.dumps({"source_id": "source-1", "action": "accept", "selected_candidate_id": "acct_1", "confidence": 0.96}, ensure_ascii=False)
        + "\n"
        + json.dumps({"source_id": "source-2", "action": "accept", "selected_candidate_id": "acct_1", "confidence": 0.97}, ensure_ascii=False)
        + "\n",
        encoding="utf-8",
    )

    summary = build_openai_resolution_candidates(
        source_workbook=workbook,
        raw_map_jsonl=raw_map,
        decisions_jsonl=decisions,
        output_dir=tmp_path / "out",
        manual_review_workbook=workbook,
    )
    rows = list(csv.DictReader(Path(summary["resolution_csv"]).open(encoding="utf-8-sig")))

    assert summary["auto_fill_openai_reranker"] == 1
    assert summary["manual_review_hard_gate"] == 1
    assert rows[0]["name"] == "홍길동"
    assert rows[0]["decision"] == "auto_fill_openai_reranker"
    assert rows[0]["chosen_account"] == "110-123-456789"
    assert rows[1]["name"] == "최검토"
    assert rows[1]["decision"] == "manual_review_hard_gate"
    assert rows[1]["chosen_account"] == ""


def test_compare_human_eval_reports_outputs_policy_vs_openai_summary(tmp_path: Path):
    policy_report = tmp_path / "policy.json"
    openai_report = tmp_path / "openai.json"
    policy_report.write_text(
        json.dumps({"summary": {"correct_positive": 2, "wrong_positive": 1}, "metrics": {"safe_selection_precision": 0.66}}),
        encoding="utf-8",
    )
    openai_report.write_text(
        json.dumps({"summary": {"correct_positive": 3, "wrong_positive": 0}, "metrics": {"safe_selection_precision": 1.0}}),
        encoding="utf-8",
    )

    summary = compare_human_eval_reports(
        reports=[("policy", policy_report), ("openai", openai_report)],
        output_dir=tmp_path / "compare",
    )
    text = Path(summary["comparison_csv"]).read_text(encoding="utf-8-sig")

    assert summary["report_count"] == 2
    assert "policy" in text
    assert "openai" in text
    assert "safe_selection_precision" in text


def test_local_oss_reranker_repairs_and_validates_json_decision():
    payload = {
        "source_id": "source-1",
        "candidates": [{"candidate_id": "acct_1", "teacher_policy_score": 12.0}],
    }
    parsed = _extract_json_object('model says {"action":"ACCEPT","selected_candidate_id":"acct_1","confidence":1.2} done')
    decision = _repair_decision(payload, parsed)

    assert decision["schema_version"] == "openai_account_rerank_decision_v1"
    assert decision["source_id"] == "source-1"
    assert decision["action"] == "accept"
    assert decision["selected_candidate_id"] == "acct_1"
    assert decision["confidence"] == 1.0
    validate_decision(payload, decision)


def test_local_cross_encoder_reranker_uses_redacted_candidate_document_and_scores():
    payload = {
        "source_id": "source-1",
        "candidates": [
            {
                "candidate_id": "acct_1",
                "account_shape": "ddd-ddd-dddddd",
                "digit_count": 12,
                "hyphen_count": 2,
                "group_count": 3,
                "repeat_count": 1,
                "teacher_policy_score": 18.0,
                "context_flags": {"has_direct_account_field_context": True},
                "risk_flags": {"looks_like_phone": False, "has_wrong_field_context": False},
            },
            {
                "candidate_id": "acct_2",
                "account_shape": "ddd-dddd-dddd",
                "digit_count": 11,
                "hyphen_count": 2,
                "group_count": 3,
                "repeat_count": 1,
                "teacher_policy_score": 9.0,
                "context_flags": {},
                "risk_flags": {"looks_like_phone": True},
            },
        ],
    }

    document = candidate_document(payload["candidates"][0])
    decision = rerank_payload_with_scores(payload, scores=[2.0, -1.0], model_id="test-model", policy_threshold=10.0)

    assert "ddd-ddd-dddddd" in document
    assert "110-123-456789" not in document
    assert decision["action"] == "accept"
    assert decision["selected_candidate_id"] == "acct_1"
    validate_decision(payload, decision)


def test_local_cross_encoder_document_includes_v2_evidence():
    candidate = {
        "candidate_id": "acct_1",
        "account_shape": "ddd-ddd-dddddd",
        "digit_count": 12,
        "hyphen_count": 2,
        "group_count": 3,
        "repeat_count": 2,
        "teacher_policy_score": 18.0,
        "context_flags": {"has_direct_account_field_context": True},
        "risk_flags": {
            "looks_like_phone": False,
            "has_wrong_field_context": False,
            "has_prompt_leakage_context": False,
        },
        "shape_features": {
            "group_lengths": [3, 3, 6],
            "pattern_family": "bank_account_like",
            "prefix_class": "known_bank_prefix",
        },
        "field_evidence": {
            "same_line_label_type": "account_number",
            "is_value_in_account_field": True,
        },
        "bank_holder_evidence": {
            "bank_name_present": True,
            "bank_name_normalized": "KB국민은행",
            "holder_match_status": "match",
        },
        "consensus_features": {
            "variant_vote_count": 2,
            "seen_in_full_ocr": True,
            "seen_in_targeted_retry": True,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "variant": "contrast",
            "prompt_id": "visible_numbers",
        },
    }

    document = candidate_document(candidate)

    assert "pattern_family=bank_account_like" in document
    assert "same_line_label_type=account_number" in document
    assert "bank_name_normalized=KB국민은행" in document
    assert "variant_vote_count=2" in document
    assert "prompt_id=visible_numbers" in document
    assert "110-123-456789" not in document


def test_v2_feature_score_accepts_strong_redacted_evidence_and_rejects_hard_risk():
    strong_candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 8.0,
        "shape_features": {"pattern_family": "bank_account_like"},
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "table_row_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {
            "bank_name_present": True,
            "holder_match_status": "match",
            "bankbook_doc_type_confidence": "high",
        },
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_full_ocr": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 2,
        },
        "risk_flags": {
            "looks_like_phone": False,
            "has_wrong_field_context": False,
            "has_prompt_leakage_context": False,
        },
    }
    risky_candidate = {
        **strong_candidate,
        "candidate_id": "acct_2",
        "teacher_policy_score": 30.0,
        "risk_flags": {"looks_like_phone": True},
    }

    assert v2_feature_score(strong_candidate) > 10.0
    assert v2_feature_score(risky_candidate) < v2_feature_score(strong_candidate)

    accepted = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [strong_candidate]},
        threshold=10.0,
        min_margin=2.0,
    )
    rejected = rerank_payload_with_v2_feature_score(
        {"source_id": "source-2", "candidates": [risky_candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert accepted["action"] == "accept"
    assert accepted["selected_candidate_id"] == "acct_1"
    assert rejected["action"] == "reject"
    assert rejected["selected_candidate_id"] is None


def test_v2_feature_score_rewards_account_kie_only_with_safe_context():
    candidate = {
        "teacher_policy_score": 7,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": False,
            "has_wrong_field_context": False,
        },
        "context_flags": {
            "has_direct_account_field_context": True,
            "has_customer_number_metadata_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "kie_evidence": {
            "field_type": "account_number",
            "confidence_bucket": "high",
        },
        "bank_holder_evidence": {
            "holder_match_status": "match",
            "bank_name_present": True,
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 1,
        },
    }

    assert v2_feature_score(candidate) >= 10


def test_v2_feature_score_blocks_customer_number_kie_even_when_numeric_shape_good():
    candidate = {
        "teacher_policy_score": 12,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": False,
            "has_wrong_field_context": False,
        },
        "context_flags": {
            "has_direct_account_field_context": False,
            "has_customer_number_metadata_context": True,
        },
        "field_evidence": {
            "is_value_in_account_field": False,
            "is_value_in_customer_number_field": True,
        },
        "kie_evidence": {
            "field_type": "customer_number",
            "confidence_bucket": "high",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": False,
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 1,
        },
    }

    assert v2_feature_score(candidate) < 10


def test_v2_feature_score_blocks_non_account_kie_fields_even_with_high_policy_score():
    candidate = {
        "teacher_policy_score": 25,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": False,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": False,
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
        },
        "kie_evidence": {
            "field_type": "phone",
            "confidence_bucket": "high",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": False,
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 1,
        },
    }

    assert v2_feature_score(candidate) < 10


def test_v2_feature_score_blocks_bank_field_kie_even_with_bank_account_shape():
    candidate = {
        "teacher_policy_score": 25,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": False,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": False,
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
        },
        "kie_evidence": {
            "backend": "paddleocr_kie",
            "field_type": "bank",
            "confidence_bucket": "high",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": True,
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 1,
        },
    }

    assert v2_feature_score(candidate) < 10


def test_v2_feature_score_blocks_ambiguous_kie_account_without_holder_match():
    candidate = {
        "teacher_policy_score": 25,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": False,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
        },
        "kie_evidence": {
            "backend": "paddleocr_kie",
            "field_type": "account_number",
            "confidence_bucket": "high",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": True,
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 7,
            "variant_vote_count": 1,
        },
    }

    assert v2_feature_score(candidate) < 10


def test_v2_feature_score_rescues_low_ambiguity_structured_retry_consensus():
    candidate = {
        "teacher_policy_score": 8,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": True,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "prompt_id": "account_structured_ko",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": True,
        },
        "consensus_features": {
            "candidate_source_count_for_person": 2,
            "same_candidate_seen_across_variants": True,
            "unique_candidate_count_for_person": 2,
            "variant_vote_count": 2,
        },
    }

    assert v2_feature_score(candidate) >= 10


def test_v2_feature_score_does_not_rescue_higher_ambiguity_structured_retry():
    candidate = {
        "teacher_policy_score": 9,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": True,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "prompt_id": "account_structured_ko",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": True,
        },
        "consensus_features": {
            "candidate_source_count_for_person": 3,
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 3,
        },
    }

    assert v2_feature_score(candidate) < 10


def test_v2_feature_score_rescues_structured_retry_only_when_repeated_four_times():
    repeated_four = {
        "teacher_policy_score": 9,
        "risk_flags": {
            "looks_like_phone": False,
            "has_prompt_leakage_context": True,
            "has_wrong_field_context": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "prompt_id": "account_structured_ko",
        },
        "bank_holder_evidence": {
            "holder_match_status": "not_present",
            "bank_name_present": True,
        },
        "consensus_features": {
            "candidate_source_count_for_person": 4,
            "same_candidate_seen_across_variants": True,
            "unique_candidate_count_for_person": 1,
            "variant_vote_count": 4,
        },
    }
    repeated_three = {
        **repeated_four,
        "consensus_features": {
            **repeated_four["consensus_features"],
            "candidate_source_count_for_person": 3,
            "variant_vote_count": 3,
        },
    }

    assert v2_feature_score(repeated_four) >= 10
    assert v2_feature_score(repeated_three) < 10


def test_v2_feature_score_recovers_visible_numbers_hyphenated_single_candidate_but_not_single_long_run():
    visible_numbers_candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 6.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "field_evidence": {
            "is_value_in_account_field": False,
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {
            "bank_name_present": False,
            "holder_match_status": "not_present",
            "bankbook_doc_type_confidence": "low",
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "same_candidate_seen_across_variants": False,
            "seen_in_full_ocr": False,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 1,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "variant": "top_crop",
            "prompt_id": "visible_numbers",
        },
        "risk_flags": {
            "looks_like_phone": False,
            "has_wrong_field_context": False,
            "has_prompt_leakage_context": False,
        },
    }
    single_long_run_candidate = {
        **visible_numbers_candidate,
        "candidate_id": "acct_2",
        "teacher_policy_score": 9.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": False,
            "is_single_long_run": True,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "variant": "original",
            "prompt_id": "account_only",
        },
    }

    recovered = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [visible_numbers_candidate]},
        threshold=10.0,
        min_margin=2.0,
    )
    rejected = rerank_payload_with_v2_feature_score(
        {"source_id": "source-2", "candidates": [single_long_run_candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert v2_feature_score(visible_numbers_candidate) >= 10.0
    assert v2_feature_score(single_long_run_candidate) < 10.0
    assert recovered["action"] == "accept"
    assert rejected["action"] == "review"


def test_v2_feature_score_recovers_visible_numbers_wrong_field_when_account_field_is_repeated():
    candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 8.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {
            "bank_name_present": True,
            "holder_match_status": "unknown",
            "bankbook_doc_type_confidence": "medium",
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "candidate_source_count_for_person": 3,
            "same_candidate_seen_across_variants": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 3,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "variant": "top_crop",
            "prompt_id": "visible_numbers",
        },
        "risk_flags": {
            "looks_like_phone": False,
            "has_wrong_field_context": True,
            "has_prompt_leakage_context": False,
        },
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert v2_feature_score(candidate) >= 10.0
    assert decision["action"] == "accept"
    assert decision["selected_candidate_id"] == "acct_1"


def test_v2_feature_score_keeps_generic_wrong_field_as_hard_risk():
    candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 30.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "field_evidence": {
            "is_value_in_account_field": False,
            "same_line_label_type": "holder",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {
            "bank_name_present": True,
            "holder_match_status": "not_present",
            "bankbook_doc_type_confidence": "medium",
        },
        "consensus_features": {
            "unique_candidate_count_for_person": 1,
            "candidate_source_count_for_person": 3,
            "same_candidate_seen_across_variants": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 3,
        },
        "source_evidence": {
            "source_kind": "targeted_retry",
            "variant": "top_crop",
            "prompt_id": "visible_numbers",
        },
        "risk_flags": {
            "looks_like_phone": False,
            "has_wrong_field_context": True,
            "has_prompt_leakage_context": False,
        },
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert decision["action"] == "reject"
    assert decision["selected_candidate_id"] is None


def test_v2_feature_score_prefers_clean_account_only_over_structured_prompt_leakage():
    clean_account_only = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 14.0,
        "shape_features": {"pattern_family": "bank_account_like", "is_single_long_run": False},
        "field_evidence": {"is_value_in_account_field": False, "is_value_in_customer_number_field": False},
        "bank_holder_evidence": {"bank_name_present": False, "holder_match_status": "not_present"},
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 2,
            "unique_candidate_count_for_person": 1,
        },
        "source_evidence": {"source_kind": "targeted_retry", "prompt_id": "account_only", "variant": "original"},
        "risk_flags": {"has_prompt_leakage_context": False, "looks_like_phone": False, "has_wrong_field_context": False},
    }
    leaked_structured = {
        "candidate_id": "acct_2",
        "teacher_policy_score": 9.0,
        "shape_features": {"pattern_family": "bank_account_like", "is_single_long_run": False},
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {"bank_name_present": True, "holder_match_status": "not_present"},
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 2,
            "unique_candidate_count_for_person": 1,
        },
        "source_evidence": {"source_kind": "targeted_retry", "prompt_id": "account_structured_ko", "variant": "original"},
        "risk_flags": {"has_prompt_leakage_context": True, "looks_like_phone": False, "has_wrong_field_context": False},
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [clean_account_only, leaked_structured]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert decision["action"] == "accept"
    assert decision["selected_candidate_id"] == "acct_1"


def test_v2_feature_score_reviews_single_long_run_from_structured_prompt_leakage():
    candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 4.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": False,
            "is_single_long_run": True,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {"bank_name_present": True, "holder_match_status": "not_present"},
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 2,
            "unique_candidate_count_for_person": 1,
        },
        "source_evidence": {"source_kind": "targeted_retry", "prompt_id": "account_structured_ko", "variant": "top_crop"},
        "risk_flags": {"has_prompt_leakage_context": True, "looks_like_phone": False, "has_wrong_field_context": False},
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert v2_feature_score(candidate) < 10.0
    assert decision["action"] == "review"


def test_v2_feature_score_reviews_structured_prompt_leakage_without_holder_or_full_ocr():
    candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 9.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {"bank_name_present": True, "holder_match_status": "not_present"},
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_full_ocr": False,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 2,
            "unique_candidate_count_for_person": 1,
        },
        "source_evidence": {"source_kind": "targeted_retry", "prompt_id": "account_structured_ko", "variant": "original"},
        "risk_flags": {"has_prompt_leakage_context": True, "looks_like_phone": False, "has_wrong_field_context": False},
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert v2_feature_score(candidate) < 10.0
    assert decision["action"] == "review"


def test_v2_feature_score_reviews_highly_ambiguous_person_without_holder_match():
    candidate = {
        "candidate_id": "acct_1",
        "teacher_policy_score": 14.0,
        "shape_features": {
            "pattern_family": "bank_account_like",
            "has_bank_style_hyphenation": True,
            "is_single_long_run": False,
        },
        "field_evidence": {
            "is_value_in_account_field": True,
            "same_line_label_type": "account_number",
            "is_value_in_customer_number_field": False,
        },
        "bank_holder_evidence": {"bank_name_present": True, "holder_match_status": "not_present"},
        "consensus_features": {
            "same_candidate_seen_across_variants": True,
            "seen_in_full_ocr": True,
            "seen_in_targeted_retry": True,
            "variant_vote_count": 3,
            "unique_candidate_count_for_person": 5,
        },
        "source_evidence": {"source_kind": "full_ocr", "prompt_id": "unknown", "variant": "unknown"},
        "risk_flags": {"has_prompt_leakage_context": False, "looks_like_phone": False, "has_wrong_field_context": False},
    }

    decision = rerank_payload_with_v2_feature_score(
        {"source_id": "source-1", "candidates": [candidate]},
        threshold=10.0,
        min_margin=2.0,
    )

    assert v2_feature_score(candidate) < 10.0
    assert decision["action"] == "review"
