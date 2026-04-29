import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook

from scripts.evaluate_human_workbook_labels import evaluate_human_workbook_labels


def build_human_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "시트"
    ws["B3"] = "1. 테스트 (4인)"
    ws["B5"] = "No."
    ws["D5"] = "성명"
    ws["J5"] = "은행명"
    ws["K5"] = "계좌번호"
    rows = [
        (6, 1, "홍길동", "국민", "110-123-456789"),
        (7, 2, "김철수", "하나", "222-333-444444"),
        (8, 3, "박영희", "농협", "555-666-777777"),
        (9, 4, "최검토", "", "888-999-000000"),
    ]
    for row_index, no, name, bank, account in rows:
        ws.cell(row_index, 2).value = no
        ws.cell(row_index, 4).value = name
        ws.cell(row_index, 10).value = bank
        ws.cell(row_index, 11).value = account

    review = wb.create_sheet("REMAINING_REVIEW")
    review.append(["group", "no", "name", "reason"])
    review.append(["1. 테스트 (4인)", 4, "최검토", "reject_holder_mismatch"])
    wb.save(path)


def test_evaluate_human_workbook_labels_splits_positive_and_review_without_raw_accounts(tmp_path: Path):
    human_workbook = tmp_path / "human.xlsx"
    resolution_csv = tmp_path / "resolution.csv"
    output_dir = tmp_path / "eval"
    build_human_workbook(human_workbook)
    resolution_csv.write_text(
        "name,decision,chosen_account,chosen_account_masked,source,candidate_files\n"
        "홍길동,auto_fill_single_candidate,110-123-456789,110-123-456789,deepseek,bank.png\n"
        "김철수,auto_fill_targeted_deepseek,222-333-999999,222-333-999999,targeted,bank.png\n"
        "박영희,no_candidate,,,,\n"
        "최검토,auto_fill_targeted_deepseek,888-999-000000,888-999-000000,targeted,bank.png\n",
        encoding="utf-8",
    )

    report = evaluate_human_workbook_labels(
        human_workbook=human_workbook,
        resolution_csv=resolution_csv,
        output_dir=output_dir,
    )
    details_text = (output_dir / "human_label_eval_details.csv").read_text(encoding="utf-8-sig")
    report_text = (output_dir / "human_label_eval.json").read_text(encoding="utf-8")
    rows = list(csv.DictReader((output_dir / "human_label_eval_details.csv").open(encoding="utf-8-sig")))

    assert report["summary"]["human_positive_count"] == 3
    assert report["summary"]["human_review_count"] == 1
    assert report["summary"]["correct_positive"] == 1
    assert report["summary"]["wrong_positive"] == 1
    assert report["summary"]["missed_positive"] == 1
    assert report["summary"]["review_false_positive"] == 1
    assert report["by_decision"]["auto_fill_targeted_deepseek"]["review_false_positive"] == 1
    assert {row["outcome"] for row in rows} == {
        "correct_positive",
        "wrong_positive",
        "missed_positive",
        "review_false_positive",
    }
    assert not re.search(r"(?<![0-9.])(?:[0-9][0-9 -]{7,22}[0-9])(?![0-9.])", details_text)
    assert "110-123-456789" not in details_text
    assert "222-333-999999" not in details_text
    assert "888-999-000000" not in report_text
