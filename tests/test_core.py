from pathlib import Path

from openpyxl import Workbook

from settlement_tool.core import (
    classify_account_candidates,
    extract_roster,
    match_files_by_name,
    normalize_text,
)


def test_extract_roster_reads_group_headers_and_people(tmp_path: Path):
    workbook_path = tmp_path / "withholding.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["B3"] = "1. 퍼실리테이터 (10인)"
    ws["B4"] = "No."
    ws["D5"] = "성명"
    ws["B6"] = 1
    ws["D6"] = "강민지"
    ws["B7"] = 2
    ws["D7"] = "임보리"
    ws["B10"] = "2. 비주얼라이터 (10인)"
    ws["B13"] = 1
    ws["D13"] = "권동현"
    wb.save(workbook_path)

    roster = extract_roster(workbook_path)

    assert [(p.group, p.no, p.name, p.row) for p in roster.people] == [
        ("1. 퍼실리테이터 (10인)", 1, "강민지", 6),
        ("1. 퍼실리테이터 (10인)", 2, "임보리", 7),
        ("2. 비주얼라이터 (10인)", 1, "권동현", 13),
    ]
    assert roster.group_counts() == {
        "1. 퍼실리테이터 (10인)": 2,
        "2. 비주얼라이터 (10인)": 1,
    }


def test_match_files_by_name_only_accepts_single_exact_name_match():
    names = ["강민지", "김영주", "이유빈"]
    files = [
        "4. 신분증 사본 업로드 (File responses)/강민지_신분증.jpg",
        "4. 신분증 사본 업로드 (File responses)/20260424 - 김영주.jpg",
        "4. 신분증 사본 업로드 (File responses)/김영주_신분증.jpeg",
        "4. 신분증 사본 업로드 (File responses)/이유빙_신분증.jpg",
    ]

    matches = match_files_by_name(names, files)

    assert matches.confirmed["강민지"].source_name == files[0]
    assert "김영주" in matches.ambiguous
    assert "이유빈" in matches.missing
    assert files[3] in matches.unmatched_files


def test_classify_account_candidates_filters_phone_numbers_and_prefers_bank_context():
    text = """
    예금주 강민지
    연락처 010-1234-5678
    신한은행 110-123-456789
    """

    result = classify_account_candidates(text)

    assert result.value == "110-123-456789"
    assert result.confidence == "high"
    assert result.candidates == ["110-123-456789"]


def test_classify_account_candidates_rejects_regional_and_toll_free_phone_numbers():
    text = """
    고객센터 054-462-1171
    해외 문의 82-2-3449-8000
    수신자 부담 080-023-0182
    """

    result = classify_account_candidates(text)

    assert result.value is None
    assert result.confidence == "none"
    assert result.candidates == []


def test_classify_account_candidates_penalizes_prompt_leakage_without_document_context():
    text = "KNOWN 계좌번호가 보이면 account_number: 110-123-456789"

    result = classify_account_candidates(text)

    assert result.value is None
    assert result.confidence == "low"
    assert result.candidates == ["110-123-456789"]


def test_classify_account_candidates_keeps_structured_bankbook_rows_high_confidence():
    text = "<tr><td>계좌번호</td><td>110-123-456789</td></tr><tr><td>예금주</td><td>홍길동</td></tr>"

    result = classify_account_candidates(text)

    assert result.value == "110-123-456789"
    assert result.confidence == "high"


def test_classify_account_candidates_defers_close_policy_reranker_candidates():
    text = """
    국민은행 계좌번호 110-123-4567
    신한은행 계좌 333-444-555555
    """

    result = classify_account_candidates(text)

    assert result.value is None
    assert result.confidence == "low"
    assert result.reason == "accepted_candidates_within_margin"
    assert result.candidates == ["110-123-4567", "333-444-555555"]


def test_classify_account_candidates_rejects_candidate_in_holder_field():
    text = "<tr><td>예금주</td><td>110-123-456789</td></tr><tr><td>계좌번호</td><td>2026.02.28</td></tr>"

    result = classify_account_candidates(text)

    assert result.value is None
    assert result.confidence == "low"
    assert result.candidates == ["110-123-456789"]


def test_normalize_text_converts_decomposed_korean_to_nfc():
    assert normalize_text("김민채") == "김민채"
