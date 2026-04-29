from pathlib import Path

from openpyxl import Workbook, load_workbook

from settlement_tool.accounts import apply_account_updates
from settlement_tool.core import AccountResult


def test_apply_account_updates_writes_only_high_confidence_results(tmp_path: Path):
    workbook_path = tmp_path / "withholding.xlsx"
    output_path = tmp_path / "updated.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["B3"] = "1. 퍼실리테이터 (10인)"
    ws["B6"] = 1
    ws["D6"] = "강민지"
    ws["B7"] = 2
    ws["D7"] = "임보리"
    wb.save(workbook_path)

    report = apply_account_updates(
        workbook_path,
        output_path,
        {
            "강민지": AccountResult("110-123-456789", "high", ["110-123-456789"], "single_high_score"),
            "임보리": AccountResult(None, "low", ["010-1234-5678"], "ambiguous_or_low_score", "chandra:hf"),
        },
    )

    updated = load_workbook(output_path, data_only=False)
    ws = updated.active
    assert ws["J6"].value == "110-123-456789"
    assert ws["J6"].number_format == "@"
    assert ws["J7"].value is None
    assert [row["status"] for row in report] == ["updated", "skipped"]
    assert report[1]["backend"] == "chandra:hf"
