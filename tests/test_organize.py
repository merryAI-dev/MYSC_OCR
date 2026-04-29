from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook

from settlement_tool.organize import build_document_plan, materialize_plan


def _write_roster(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["B3"] = "1. 퍼실리테이터 (10인)"
    ws["B6"] = 1
    ws["D6"] = "강민지"
    ws["B7"] = 2
    ws["D7"] = "임보리"
    wb.save(path)


def _write_zip(path: Path, members: dict[str, bytes]) -> None:
    with ZipFile(path, "w") as zf:
        for name, payload in members.items():
            zf.writestr(name, payload)


def test_build_document_plan_uses_group_and_standard_filename(tmp_path: Path):
    roster_path = tmp_path / "withholding.xlsx"
    _write_roster(roster_path)
    payment_zip = tmp_path / "payments.zip"
    id_zip = tmp_path / "ids.zip"
    bank_zip = tmp_path / "banks.zip"
    _write_zip(
        payment_zip,
        {
            "root/1. 퍼실리테이터 (10인)/1-1. 강민지_비용지급확인서.pdf": b"pay",
            "root/1. 퍼실리테이터 (10인)/2-1. 임보리_비용지급확인서.pdf": b"pay2",
        },
    )
    _write_zip(id_zip, {"ids/강민지_신분증.jpg": b"id"})
    _write_zip(bank_zip, {"banks/강민지_통장사본.png": b"bank"})

    plan = build_document_plan(roster_path, payment_zip, id_zip, bank_zip, overrides={})

    gang = [item for item in plan.items if item.name == "강민지"]
    assert {item.doc_type for item in gang if item.status == "confirmed"} == {
        "비용지급확인서",
        "신분증",
        "통장사본",
    }
    assert gang[0].group == "1. 퍼실리테이터 (10인)"
    assert "임보리" in {item.name for item in plan.items if item.status == "missing"}


def test_materialize_plan_copies_confirmed_files_to_standard_names(tmp_path: Path):
    roster_path = tmp_path / "withholding.xlsx"
    _write_roster(roster_path)
    payment_zip = tmp_path / "payments.zip"
    id_zip = tmp_path / "ids.zip"
    bank_zip = tmp_path / "banks.zip"
    _write_zip(payment_zip, {"root/1. 퍼실리테이터 (10인)/1-1. 강민지_비용지급확인서.pdf": b"pay"})
    _write_zip(id_zip, {"ids/강민지_신분증.jpg": b"id"})
    _write_zip(bank_zip, {"banks/강민지_통장사본.png": b"bank"})
    plan = build_document_plan(roster_path, payment_zip, id_zip, bank_zip, overrides={})

    materialize_plan(plan, tmp_path / "out", dry_run=False)

    assert (tmp_path / "out" / "1. 퍼실리테이터 (10인)" / "강민지_비용지급확인서.pdf").read_bytes() == b"pay"
    assert (tmp_path / "out" / "1. 퍼실리테이터 (10인)" / "강민지_신분증.jpg").read_bytes() == b"id"
    assert (tmp_path / "out" / "1. 퍼실리테이터 (10인)" / "강민지_통장사본.png").read_bytes() == b"bank"
